# GrassCRM Backend — main.py v1.2

from dotenv import load_dotenv
load_dotenv()

import os
import secrets
import threading
import time as _time
import shutil
import importlib.util
from pathlib import Path
from datetime import datetime, date, timedelta
from contextlib import asynccontextmanager
from typing import Optional, List
from functools import lru_cache

from fastapi import FastAPI, Depends, HTTPException, status, Request, Header, UploadFile, File as FastAPIFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, FileResponse
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy import (
    create_engine, Column, Integer, String, Float, Date, DateTime, 
    Boolean, ForeignKey, Text, text, MetaData, extract, Double, func
)
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, Session as DBSession, joinedload
from pydantic import BaseModel, Field, ConfigDict
import httpx
from jose import jwt, JWTError

# ── 1. КОНФИГ ─────────────────────────────────────────────────────────────────
DATABASE_URL   = os.getenv("DATABASE_URL")
AUTH0_DOMAIN   = os.getenv("AUTH0_DOMAIN")
AUTH0_AUDIENCE = os.getenv("AUTH0_AUDIENCE")
CLIENT_ID      = os.getenv("AUTH0_CLIENT_ID")
CLIENT_SECRET  = os.getenv("AUTH0_CLIENT_SECRET")
APP_BASE_URL   = os.getenv("APP_BASE_URL", "https://crmpokos.ru").rstrip("/")
CALLBACK_URL   = f"{APP_BASE_URL}/api/auth/callback"
SESSION_SECRET = os.getenv("SESSION_SECRET", secrets.token_hex(32))
INTERNAL_API_KEY = os.getenv("INTERNAL_API_KEY")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
OPENAI_ACCESS_ID = os.getenv("OPENAI_ACCESS_ID") or os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
ROLE_CLAIM     = "https://grass-crm/role"
CACHE_TTL      = 300
TAX_RATE       = float(os.getenv("TAX_RATE", "0.04")) # 4% УСН "Доходы" для самозанятых
UPLOAD_DIR     = Path(os.getenv("UPLOAD_DIR", "/var/www/crm/GCRM-2/uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_FILE_SIZE  = 20 * 1024 * 1024  # 20 MB

# ── 2. КЭШ И УТИЛИТЫ ──────────────────────────────────────────────────────────
class _Cache:
    def __init__(self, ttl: int):
        self._ttl, self._data, self._ts, self._lock = ttl, {}, {}, threading.Lock()
    def get(self, key: str):
        if key not in self._data or _time.monotonic() - self._ts[key] > self._ttl: return None
        return self._data[key]
    def set(self, key: str, value):
        with self._lock: self._data[key], self._ts[key] = value, _time.monotonic()
    def invalidate(self, *keys):
        with self._lock:
            if not keys or "all" in keys: self._data.clear(); self._ts.clear(); return
            for k in keys:
                to_remove = [ek for ek in self._data if ek == k or ek.startswith(f"{k}:")]
                for ek in to_remove: 
                    if ek in self._data: del self._data[ek]
                    if ek in self._ts: del self._ts[ek]
_cache = _Cache(ttl=CACHE_TTL)

# ── 3. БАЗА ДАННЫХ ────────────────────────────────────────────────────────────
Base = declarative_base()
engine = create_engine(DATABASE_URL, client_encoding='utf8')
SessionFactory = sessionmaker(bind=engine, autoflush=False)

class User(Base): __tablename__ = "users"; id,username,name,email = Column(String, primary_key=True),Column(String),Column(String),Column(String); role=Column(String, default="User"); telegram_id=Column(String(50), unique=True, index=True, nullable=True); last_login=Column(DateTime, nullable=True)
class Service(Base): __tablename__ = "services"; id,name,price,unit = Column(Integer,primary_key=True),Column(String(200),nullable=False),Column(Float,default=0.0),Column(String(50),default="шт"); min_volume=Column(Float,default=1.0); notes=Column(Text)
class DealService(Base): __tablename__ = "deal_services"; id,deal_id,service_id,quantity,price_at_moment = Column(Integer,primary_key=True),Column(Integer,ForeignKey("deals.id",ondelete="CASCADE")),Column(Integer,ForeignKey("services.id",ondelete="RESTRICT")),Column(Float,default=1.0),Column(Float,nullable=False); service = relationship("Service")
class Stage(Base): __tablename__ = "stages"; id,name,order,type,is_final,color = Column(Integer,primary_key=True),Column(String(100),nullable=False,unique=True),Column(Integer,default=0),Column(String(50),default="regular"),Column(Boolean,default=False),Column(String(20),default="#6B7280"); deals = relationship("Deal", back_populates="stage")
class Contact(Base):
    __tablename__ = "contacts"
    id = Column(Integer, primary_key=True)
    name = Column(String(200), nullable=False)
    phone = Column(String(50), unique=True, index=True)
    source = Column(String(100))
    telegram_id = Column(String(50), unique=True, index=True, nullable=True)
    telegram_username = Column(String(100), nullable=True)
    addresses = Column(Text, nullable=True)  # JSON-список адресов
    settlement = Column(String(200), nullable=True)   # Название поселка/СНТ
    plot_area = Column(Float, nullable=True)           # Количество соток
    deals = relationship("Deal", back_populates="contact")
class Deal(Base): 
    __tablename__ = "deals"
    id=Column(Integer,primary_key=True)
    contact_id=Column(Integer,ForeignKey("contacts.id"),nullable=False)
    stage_id=Column(Integer,ForeignKey("stages.id"))
    title=Column(String(200),nullable=False)
    total=Column(Float,default=0.0)
    notes=Column(Text,default="")
    created_at=Column(DateTime,default=datetime.utcnow)
    deal_date=Column(DateTime)
    closed_at=Column(DateTime)
    is_repeat=Column(Boolean,default=False)
    manager=Column(String(200))
    address=Column(Text)
    tax_rate = Column(Float, default=4.0, nullable=False)
    tax_included = Column(Boolean, default=True, nullable=False)
    discount = Column(Float, default=0.0, nullable=False)
    discount_type = Column(String(10), default="percent", nullable=False)
    repeat_interval_days = Column(Integer, nullable=True)   # NULL = не повторяется
    next_repeat_date = Column(Date, nullable=True)          # дата следующего выезда

    contact=relationship("Contact",back_populates="deals")
    stage=relationship("Stage",back_populates="deals")
    services=relationship("DealService",cascade="all, delete-orphan",passive_deletes=True)

class Task(Base): __tablename__="tasks"; id,title,description,is_done=Column(Integer,primary_key=True),Column(String,nullable=False),Column(Text),Column(Boolean,default=False); due_date,assignee,priority,status=Column(Date),Column(String),Column(String,default="Обычный"),Column(String,default="Открыта"); contact_id=Column(Integer,ForeignKey("contacts.id",ondelete="SET NULL"),nullable=True); deal_id=Column(Integer,ForeignKey("deals.id",ondelete="SET NULL"),nullable=True); contact=relationship("Contact"); deal=relationship("Deal")

class Interaction(Base):
    __tablename__ = "interactions"
    id = Column(Integer, primary_key=True)
    contact_id = Column(Integer, ForeignKey("contacts.id", ondelete="CASCADE"), nullable=False)
    type = Column(String(50), nullable=False, default="note")  # call, email, meeting, note
    text = Column(Text, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    user_id = Column(String)
    user_name = Column(String)
    contact = relationship("Contact")

class DealComment(Base):
    __tablename__ = "deal_comments"
    id = Column(Integer, primary_key=True)
    deal_id = Column(Integer, ForeignKey("deals.id", ondelete="CASCADE"), nullable=False)
    text = Column(Text, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    user_id = Column(String)
    user_name = Column(String)
    deal = relationship("Deal")

class CRMFile(Base):
    __tablename__ = "crm_files"
    id = Column(Integer, primary_key=True)
    filename = Column(String(300), nullable=False)       # оригинальное имя файла
    stored_name = Column(String(300), nullable=False)    # имя на диске (уникальное)
    size = Column(Integer, default=0)                    # байты
    mime_type = Column(String(100))
    contact_id = Column(Integer, ForeignKey("contacts.id", ondelete="SET NULL"), nullable=True)
    deal_id = Column(Integer, ForeignKey("deals.id", ondelete="SET NULL"), nullable=True)
    uploaded_by = Column(String)
    uploaded_by_name = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)
    file_kind = Column(String(20), nullable=True)  # before | after | general
    contact = relationship("Contact")
    deal = relationship("Deal")

class ExpenseCategory(Base): __tablename__="expense_categories"; id,name=Column(Integer,primary_key=True),Column(String(100),nullable=False,unique=True); expenses=relationship("Expense",back_populates="category")
class Consumable(Base): __tablename__="consumables"; id,name,unit=Column(Integer,primary_key=True),Column(String(200),nullable=False,unique=True),Column(String(50),default="шт"); stock_quantity,notes,price=Column(Float,default=0.0),Column(Text),Column(Float,default=0.0)
class MaintenanceConsumable(Base): __tablename__="maintenance_consumables"; id=Column(Integer,primary_key=True); maintenance_id=Column(Integer,ForeignKey("equipment_maintenance.id",ondelete="CASCADE"),nullable=False); consumable_id=Column(Integer,ForeignKey("consumables.id",ondelete="RESTRICT"),nullable=False); quantity=Column(Float,nullable=False); price_at_moment=Column(Float,nullable=False); consumable=relationship("Consumable"); maintenance_record=relationship("EquipmentMaintenance",back_populates="consumables_used")
class EquipmentMaintenance(Base): __tablename__ = "equipment_maintenance"; id=Column(Integer,primary_key=True); equipment_id=Column(Integer,ForeignKey("equipment.id",ondelete="CASCADE"),nullable=False); date=Column(Date,nullable=False); work_description=Column(Text,nullable=False); cost=Column(Float); notes=Column(Text); equipment = relationship("Equipment", back_populates="maintenance_records"); consumables_used = relationship("MaintenanceConsumable", back_populates="maintenance_record", cascade="all, delete-orphan")
class Equipment(Base): __tablename__="equipment"; id,name,model,serial=Column(Integer,primary_key=True),Column(String(200),nullable=False),Column(String(200),default=""),Column(String(100)); purchase_date,purchase_cost=Column(Date),Column(Double,default=0.0); status,notes=Column(String(50),default="active"),Column(Text); engine_hours=Column(Double,default=0.0); fuel_norm=Column(Double,default=0.0); last_maintenance_date=Column(Date); next_maintenance_date=Column(Date); expenses=relationship("Expense",back_populates="equipment"); maintenance_records = relationship("EquipmentMaintenance", back_populates="equipment", cascade="all, delete-orphan")
class Expense(Base): __tablename__="expenses"; id,date,name,amount=Column(Integer,primary_key=True),Column(Date,nullable=False,default=date.today),Column(String(300),nullable=False),Column(Float,nullable=False); category_id,equipment_id=Column(Integer,ForeignKey("expense_categories.id")),Column(Integer,ForeignKey("equipment.id")); category=relationship("ExpenseCategory",back_populates="expenses"); equipment=relationship("Equipment",back_populates="expenses")

class TaxPayment(Base):
    __tablename__ = "tax_payments"
    id = Column(Integer, primary_key=True)
    amount = Column(Float, nullable=False)
    date = Column(Date, nullable=False, default=date.today)
    note = Column(Text)
    year = Column(Integer, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class DailyPhrase(Base):
    __tablename__ = "daily_phrases"
    id = Column(Integer, primary_key=True)
    phrase = Column(Text, nullable=False)
    category = Column(String(50), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class AiMemory(Base):
    """Долгосрочная память AI-ассистента."""
    __tablename__ = "ai_memory"
    id          = Column(Integer, primary_key=True)
    chat_id     = Column(String(50), nullable=False, index=True)
    role        = Column(String(20), nullable=False, default="user")   # user|assistant|fact
    content     = Column(Text, nullable=False)
    memory_type = Column(String(20), default="message")  # message|fact|preference
    importance  = Column(Integer, default=1)              # 1=обычное 2=важное 3=критическое
    metadata_   = Column("metadata", JSONB, default=dict)
    created_at  = Column(DateTime, default=datetime.utcnow)
    expires_at  = Column(DateTime, nullable=True)


class AiUsageLog(Base):
    __tablename__ = "ai_usage_log"
    id = Column(Integer, primary_key=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    prompt_tokens = Column(Integer, default=0)
    completion_tokens = Column(Integer, default=0)
    total_tokens = Column(Integer, default=0)
    source = Column(String(50), default="crm")

class BotFaq(Base):
    __tablename__ = "bot_faq"
    id = Column(Integer, primary_key=True)
    intent = Column(String(100), nullable=False)
    question_example = Column(Text)
    answer = Column(Text, nullable=False)
    priority = Column(Integer, default=10)
    active = Column(Boolean, default=True)

class Note(Base):
    __tablename__ = "notes"
    id          = Column(String, primary_key=True)
    user_id     = Column(String, nullable=False, index=True)
    title       = Column(Text, default="")
    body        = Column(Text, default="")
    color       = Column(String(50), default="")
    pinned      = Column(Boolean, default=False)
    label       = Column(String(100), default="")
    checklist   = Column(JSONB, default=list)
    created_at  = Column(DateTime, default=datetime.utcnow)
    updated_at  = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

def init_db_structure(): Base.metadata.create_all(engine)
def seed_initial_data(s: DBSession):
    if s.query(Stage).count()==0: s.add_all([Stage(**d) for d in [{"name":"Согласовать","order":1,"color":"#3B82F6"},{"name":"Ожидание","order":2,"color":"#F59E0B"},{"name":"В работе","order":3,"color":"#EC4899"},{"name":"Успешно","order":4,"color":"#10B981","is_final":True},{"name":"Провалена","order":5,"color":"#EF4444","is_final":True}]])
    if s.query(ExpenseCategory).count()==0: s.add_all([ExpenseCategory(name=n) for n in ["Техника","Топливо","Расходники","Реклама","Запчасти","Прочее"]])
    s.commit()

# ── 4. HELPERS ────────────────────────────────────────────────────────────────
def update_equipment_last_maintenance(db: DBSession, equipment_id: int):
    equipment_item = db.query(Equipment).filter(Equipment.id == equipment_id).first();
    if not equipment_item: return
    latest_maintenance_date = db.query(func.max(EquipmentMaintenance.date)).filter(EquipmentMaintenance.equipment_id == equipment_id).scalar()
    equipment_item.last_maintenance_date = latest_maintenance_date; db.commit()
    _cache.invalidate("equipment")

# ── 5. PYDANTIC MODELS ───────────────────────────────────────────────────────
class DealServiceItem(BaseModel): service_id: int; quantity: float
class DealCreate(BaseModel): 
    title:str; 
    stage_id:int; 
    contact_id:Optional[int]=None; 
    new_contact_name:Optional[str]=None; 
    manager:Optional[str]=None; 
    services:List[DealServiceItem]=[]
    tax_rate: Optional[float] = 4.0
    tax_included: Optional[bool] = True
    discount: Optional[float] = 0.0
    discount_type: Optional[str] = "percent"
    work_date: Optional[str] = None
    work_time: Optional[str] = None
    address: Optional[str] = None
    repeat_interval_days: Optional[int] = None
    next_repeat_date: Optional[str] = None

class DealUpdate(BaseModel): 
    title:Optional[str]=None; 
    stage_id:Optional[int]=None; 
    contact_id:Optional[int]=None; 
    new_contact_name:Optional[str]=None; 
    manager:Optional[str]=None; 
    services:Optional[List[DealServiceItem]]=None
    tax_rate: Optional[float] = None
    tax_included: Optional[bool] = None
    discount: Optional[float] = None
    discount_type: Optional[str] = None
    work_date: Optional[str] = None
    work_time: Optional[str] = None
    address: Optional[str] = None
    repeat_interval_days: Optional[int] = None
    next_repeat_date: Optional[str] = None

class TaskCreate(BaseModel): title: str; description: Optional[str]=None; due_date: Optional[date]=None; priority: Optional[str]="Обычный"; status: Optional[str]="Открыта"; assignee: Optional[str]=None; contact_id: Optional[int]=None; deal_id: Optional[int]=None
class TaskUpdate(BaseModel): title: Optional[str]=None; description: Optional[str]=None; due_date: Optional[date]=None; priority: Optional[str]=None; status: Optional[str]=None; assignee: Optional[str]=None; is_done: Optional[bool]=None; contact_id: Optional[int]=None; deal_id: Optional[int]=None
class ContactCreate(BaseModel): name: str = Field(..., min_length=1); phone: Optional[str] = None; source: Optional[str] = None; telegram_id: Optional[str] = None; telegram_username: Optional[str] = None; addresses: Optional[List[str]] = None; settlement: Optional[str] = None; plot_area: Optional[float] = None
class ContactUpdate(BaseModel): name: Optional[str] = Field(None, min_length=1); phone: Optional[str] = None; source: Optional[str] = None; telegram_id: Optional[str] = None; telegram_username: Optional[str] = None; addresses: Optional[List[str]] = None; settlement: Optional[str] = None; plot_area: Optional[float] = None
class ServiceCreate(BaseModel): name: str = Field(...,min_length=1); price: float; unit: str; min_volume: Optional[float]=1.0; notes: Optional[str]=None
class ServiceUpdate(BaseModel): name: Optional[str]=Field(None,min_length=1); price: Optional[float]=None; unit: Optional[str]=None; min_volume: Optional[float]=None; notes: Optional[str]=None
class EquipmentCreate(BaseModel): name: str; model: Optional[str]=None; serial: Optional[str]=None; purchase_date: Optional[date]=None; purchase_cost: Optional[float]=None; status: Optional[str]='active'; notes: Optional[str]=None; engine_hours: Optional[float]=None; fuel_norm: Optional[float]=None; last_maintenance_date: Optional[date]=None; next_maintenance_date: Optional[date]=None
class EquipmentUpdate(BaseModel): name: Optional[str]=None; model: Optional[str]=None; serial: Optional[str]=None; purchase_date: Optional[date]=None; purchase_cost: Optional[float]=None; status: Optional[str]=None; notes: Optional[str]=None; engine_hours: Optional[float]=None; fuel_norm: Optional[float]=None; last_maintenance_date: Optional[date]=None; next_maintenance_date: Optional[date]=None
class ConsumableCreate(BaseModel): name: str; unit: Optional[str] = 'шт'; stock_quantity: Optional[float] = 0.0; price: Optional[float] = 0.0; notes: Optional[str] = None
class ConsumableUpdate(BaseModel): name: Optional[str] = None; unit: Optional[str] = None; stock_quantity: Optional[float] = None; price: Optional[float] = None; notes: Optional[str] = None
class MaintenanceConsumableItem(BaseModel): consumable_id: int; quantity: float
class MaintenanceCreate(BaseModel): equipment_id: int; date: date; work_description: str; notes: Optional[str]=None; consumables: List[MaintenanceConsumableItem] = []
class MaintenanceUpdate(BaseModel): date: Optional[date]=None; work_description: Optional[str]=None; notes: Optional[str]=None; consumables: Optional[List[MaintenanceConsumableItem]] = None
class ExpenseCreate(BaseModel): name: str; amount: float; date: str; category: str
class ExpenseUpdate(BaseModel):
    name: Optional[str] = None
    amount: Optional[float] = None
    date: Optional[str] = None   # accept string, convert manually
    category: Optional[str] = None
class TaxPaymentCreate(BaseModel): amount: float; date: date; note: Optional[str] = None; year: int
class InteractionCreate(BaseModel): type: str = "note"; text: str = Field(..., min_length=1)
class DealCommentCreate(BaseModel): text: str = Field(..., min_length=1)
class UserTelegramUpdate(BaseModel): telegram_id: Optional[str] = None
class ServiceAIAgentRequest(BaseModel):
    prompt: str = Field(..., min_length=3, max_length=5000)
    year: Optional[int] = None
    base_url: Optional[str] = None
    access_id: Optional[str] = None
    model: Optional[str] = None
    # Контекст записи (для AI-кнопок внутри сделки/задачи/расхода)
    context_type: Optional[str] = None   # 'deal' | 'task' | 'expense' | 'contact'
    context_id: Optional[int] = None     # id записи


class AIActionRequest(BaseModel):
    """Запрос на выполнение AI-действия напрямую."""
    action: str
    data: dict = {}
    source: Optional[str] = 'crm'


# --- Response Models to prevent serialization cycles ---
class EquipmentForMaintResponse(BaseModel):
    id: int; name: str
    model_config = ConfigDict(from_attributes=True)

class MaintenanceForListResponse(BaseModel):
    id: int; date: date; work_description: str; cost: Optional[float]; equipment_id: int
    equipment: EquipmentForMaintResponse
    model_config = ConfigDict(from_attributes=True)

class ConsumableForMaintResponse(BaseModel):
    id: int; name: str; unit: Optional[str]; stock_quantity: Optional[float] = 0.0
    model_config = ConfigDict(from_attributes=True)

class MaintConsumableForDetailResponse(BaseModel):
    quantity: float; price_at_moment: float
    consumable: ConsumableForMaintResponse
    model_config = ConfigDict(from_attributes=True)

class MaintenanceDetailResponse(BaseModel):
    id: int; equipment_id: int; date: date; work_description: str; notes: Optional[str]; cost: Optional[float]
    consumables_used: List[MaintConsumableForDetailResponse]
    model_config = ConfigDict(from_attributes=True)

class EquipmentResponse(BaseModel):
    id:int; name:str; model:Optional[str]; serial:Optional[str]; purchase_date:Optional[date]; purchase_cost:Optional[float]; status:Optional[str]; notes:Optional[str]; engine_hours:Optional[float]; fuel_norm:Optional[float]; last_maintenance_date:Optional[date]; next_maintenance_date:Optional[date]
    model_config = ConfigDict(from_attributes=True)

class NoteCreate(BaseModel):
    id: Optional[str] = None
    title: str = ""
    body: str = ""
    color: str = ""
    pinned: bool = False
    label: str = ""
    checklist: list = []

class NoteUpdate(BaseModel):
    title: Optional[str] = None
    body: Optional[str] = None
    color: Optional[str] = None
    pinned: Optional[bool] = None
    label: Optional[str] = None
    checklist: Optional[list] = None

class BudgetResponse(BaseModel):
    id: int
    year: int
    period: str
    name: str
    planned_revenue: Optional[float] = 0.0
    planned_expenses: Optional[float] = 0.0
    notes: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)

# ── 6. FASTAPI APP ────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("App starting (v12.1)...",flush=True)
    init_db_structure()
    with SessionFactory() as db: seed_initial_data(db)
    _ensure_budget_table()
    _ensure_users_telegram_column()
    _ensure_files_kind_column()
    _ensure_contacts_telegram_columns()
    _ensure_deals_discount_type()
    _ensure_repeat_columns()
    _ensure_user_last_login()
    threading.Thread(target=_repeat_deals_worker, daemon=True).start()
    yield
    print("App shutting down.",flush=True)

app = FastAPI(title="GrassCRM API", version="12.2.0", lifespan=lifespan)
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, https_only=True, same_site="lax")
app.add_middleware(CORSMiddleware, allow_origins=[APP_BASE_URL], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

@lru_cache(maxsize=1)
def get_jwks(): return httpx.get(f"https://{AUTH0_DOMAIN}/.well-known/jwks.json", timeout=10).raise_for_status().json()
def get_db(): db = SessionFactory(); yield db; db.close()

def get_current_user(req: Request, x_internal_api_key: Optional[str] = Header(None, alias="X-Internal-API-Key")):
    # API Key authentication for internal services (like the Telegram bot)
    if INTERNAL_API_KEY and x_internal_api_key == INTERNAL_API_KEY:
        # This is a trusted internal service. We grant it a synthetic admin user profile.
        return {"sub": "internal-bot", "name": "Telegram Bot", "role": "Admin"}

    # Standard session-based authentication for web users
    if user := req.session.get("user"):
        return user

    # If neither authentication method is successful, deny access.
    raise HTTPException(status_code=401, detail="Not authenticated")

def is_admin(user: dict) -> bool:
    return (user.get("role") or "").strip().lower() == "admin"

def require_admin(user: dict = Depends(get_current_user)):
    if not is_admin(user):
        raise HTTPException(403, "Доступ запрещён: требуется роль Admin")
    return user

# ── 7. ЭНДПОИНТЫ AUTH ─────────────────────────────────────────────────────────
@app.get("/api/auth/login")
def login(req:Request): state=secrets.token_urlsafe(16); req.session["oauth_state"]=state; return RedirectResponse(f"https://{AUTH0_DOMAIN}/authorize?response_type=code&client_id={CLIENT_ID}&redirect_uri={CALLBACK_URL}&scope=openid%20profile%20email&audience={AUTH0_AUDIENCE}&state={state}")
@app.get("/api/auth/callback")
def callback(req:Request, code:str=None, state:str=None, error:str=None):
    if error: return RedirectResponse(f"/?auth_error={error}")
    if not code or state!=req.session.pop("oauth_state",None): raise HTTPException(400,"Invalid state or no code")
    with httpx.Client() as c:
        tokens=c.post(f"https://{AUTH0_DOMAIN}/oauth/token",json={"grant_type":"authorization_code","client_id":CLIENT_ID,"client_secret":CLIENT_SECRET,"code":code,"redirect_uri":CALLBACK_URL}).raise_for_status().json()
        profile=c.get(f"https://{AUTH0_DOMAIN}/userinfo",headers={"Authorization":f"Bearer {tokens['access_token']}"}).raise_for_status().json()
    user_id=profile.get("sub"); user_name=profile.get("name") or profile.get("nickname") or user_id
    with SessionFactory() as db:
        user=db.query(User).filter(User.id==user_id).first()
        if not user: db.add(User(id=user_id,name=user_name,email=profile.get("email"),last_login=datetime.utcnow()))
        else:
            user.name,user.email = user_name,profile.get("email")
            user.last_login = datetime.utcnow()
        db.commit()
    # Роль берём из БД (задаётся вручную в Supabase: Admin / User)
    with SessionFactory() as _db:
        _u = _db.query(User).filter(User.id == user_id).first()
        user_role = (_u.role or "User") if _u else "User"
    req.session["user"] = {"sub": user_id, "name": user_name, "role": user_role}
    return RedirectResponse("/")
@app.get("/api/auth/logout")
def logout(req:Request): req.session.clear(); return RedirectResponse(f"https://{AUTH0_DOMAIN}/v2/logout?client_id={CLIENT_ID}&returnTo={APP_BASE_URL}")

# ── 8. ЭНДПОИНТЫ API ──────────────────────────────────────────────────────────
@app.get("/api/me")
def get_me(user:dict=Depends(get_current_user)): return user
@app.get("/api/users")
def get_users(db:DBSession=Depends(get_db),_=Depends(get_current_user)):
    users = db.query(User).order_by(User.name).all()
    return [{"id":u.id,"username":u.username,"name":u.name,"email":u.email,"role":u.role,
             "telegram_id":u.telegram_id,
             "last_login":u.last_login.isoformat() if u.last_login else None} for u in users]

@app.get("/api/users/by-telegram/{telegram_id}")
def get_user_by_telegram(telegram_id: str, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    user = db.query(User).filter(User.telegram_id == telegram_id).first()
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    return user

@app.patch("/api/users/{user_id}/telegram")
def set_user_telegram_id(user_id: str, data: UserTelegramUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "Пользователь не найден")

    new_tid = (data.telegram_id or "").strip() or None
    if new_tid:
        conflict = db.query(User).filter(User.telegram_id == new_tid, User.id != user_id).first()
        if conflict:
            raise HTTPException(400, "Этот Telegram ID уже привязан к другому пользователю")
    user.telegram_id = new_tid
    db.commit()
    db.refresh(user)
    return user
@app.get("/api/stages")
def get_stages(db:DBSession=Depends(get_db),_=Depends(get_current_user)): return db.query(Stage).order_by(Stage.order).all()
@app.post("/api/cache/invalidate")
def invalidate_cache(_=Depends(get_current_user)): _cache.invalidate("all"); return {"status":"ok"}

# --- SERVICES ---
@app.get("/api/services")
def get_services(db:DBSession=Depends(get_db),_=Depends(get_current_user)): return db.query(Service).order_by(Service.id).all()
@app.post("/api/services", status_code=201)
def create_service(data: ServiceCreate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    new_item = Service(**data.model_dump()); db.add(new_item); db.commit(); db.refresh(new_item)
    _cache.invalidate("services"); return new_item
@app.patch("/api/services/{service_id}")
def update_service(service_id: int, data: ServiceUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = db.query(Service).filter(Service.id == service_id).first()
    if not item: raise HTTPException(404, "Услуга не найдена")
    for key, value in data.model_dump(exclude_unset=True).items(): setattr(item, key, value)
    db.commit(); db.refresh(item)
    _cache.invalidate("services", "deals"); return item
@app.delete("/api/services/{service_id}", status_code=204)
def delete_service(service_id: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    if db.query(DealService).filter(DealService.service_id == service_id).count() > 0:
        raise HTTPException(400, "Нельзя удалить услугу, которая используется в сделках.")
    item = db.query(Service).filter(Service.id == service_id).first()
    if item: db.delete(item); db.commit(); _cache.invalidate("services")
    return None

# --- DEALS ---
@app.get("/api/deals")
def get_deals(year:Optional[int]=None,db:DBSession=Depends(get_db),user:dict=Depends(get_current_user)):
    q=db.query(Deal).options(joinedload(Deal.contact),joinedload(Deal.stage)).order_by(Deal.created_at.desc())
    if year: q = q.filter(extract("year", Deal.deal_date) == year)
    if not is_admin(user): q = q.filter(Deal.manager == user["name"])  # менеджер видит только свои
    deals_list=[{"id":d.id,"title":d.title or "","total":d.total or 0.0,"client":d.contact.name if d.contact else "","contact_id":d.contact_id,"stage":d.stage.name if d.stage else "","stage_id":d.stage_id,"created_at":(d.created_at or datetime.utcnow()).isoformat()} for d in q.all()]
    return {"deals": deals_list}

@app.post("/api/deals", status_code=201)
def create_deal(deal_data: DealCreate, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    contact_id = deal_data.contact_id
    if deal_data.new_contact_name:
        new_contact = Contact(name=deal_data.new_contact_name)
        db.add(new_contact)
        db.flush()
        db.refresh(new_contact)
        contact_id = new_contact.id
    if not contact_id:
        raise HTTPException(400, "Не указан клиент")

    subtotal = 0
    service_items_for_db = []
    for item in deal_data.services:
        service = db.query(Service).filter(Service.id == item.service_id).first()
        if not service: continue
        price = service.price or 0
        subtotal += price * item.quantity
        service_items_for_db.append(DealService(service_id=service.id, quantity=item.quantity, price_at_moment=price))

    discount_val = deal_data.discount or 0
    discount_type = deal_data.discount_type or "percent"
    tax_rate_percent = deal_data.tax_rate or 0
    
    if discount_type == "fixed":
        discount_amount = min(discount_val, subtotal)
    else:
        discount_amount = subtotal * (discount_val / 100.0)
    subtotal_after_discount = subtotal - discount_amount
    
    # Налог включён: сумма = subtotal, налог отображается отдельно но не прибавляется
    # Налог сверху: сумма = subtotal + налог
    tax_amount = subtotal_after_discount * (tax_rate_percent / 100.0)
    if deal_data.tax_included:
        final_total = subtotal_after_discount
    else:
        final_total = subtotal_after_discount + tax_amount

    new_deal = Deal(
        title=deal_data.title,
        stage_id=deal_data.stage_id,
        contact_id=contact_id,
        deal_date=datetime.utcnow(),
        manager=deal_data.manager,
        services=service_items_for_db,
        total=round(final_total, 2),
        discount=discount_val,
        discount_type=discount_type,
        tax_rate=tax_rate_percent,
        tax_included=deal_data.tax_included,
        address=deal_data.address or None
    )
    # Собираем дату+время выезда из отдельных строк
    if deal_data.work_date:
        try:
            dt_str = deal_data.work_date
            if deal_data.work_time:
                dt_str += f"T{deal_data.work_time}"
            new_deal.deal_date = datetime.fromisoformat(dt_str)
        except Exception:
            pass
    # Повтор
    if deal_data.repeat_interval_days and deal_data.next_repeat_date:
        try:
            from datetime import date as _d
            new_deal.repeat_interval_days = deal_data.repeat_interval_days
            new_deal.next_repeat_date = _d.fromisoformat(str(deal_data.next_repeat_date)[:10])
        except Exception:
            pass
    db.add(new_deal)
    db.commit()
    db.refresh(new_deal)
    _cache.invalidate("deals", "years")

    return {"status": "ok", "id": new_deal.id}

@app.get("/api/deals/{deal_id}")
def get_deal_details(deal_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    deal = db.query(Deal).options(
        joinedload(Deal.contact), 
        joinedload(Deal.services).joinedload(DealService.service)
    ).filter(Deal.id == deal_id).first()

    if not deal:
        raise HTTPException(404, "Сделка не найдена")

    services_list = []
    for ds in deal.services:
        service_info = {"quantity": ds.quantity, "price_at_moment": ds.price_at_moment}
        if ds.service:
            service_info["service"] = {"id": ds.service.id, "name": ds.service.name, "price": ds.service.price, "unit": ds.service.unit}
        else:
            service_info["service"] = {"id": -1, "name": "[Удаленная услуга]", "price": ds.price_at_moment, "unit": "?"}
        services_list.append(service_info)
    
    return {
        "id": deal.id,
        "title": deal.title,
        "total": deal.total,
        "stage_id": deal.stage_id,
        "manager": deal.manager,
        "contact": {"id": deal.contact.id, "name": deal.contact.name} if deal.contact else None,
        "services": services_list,
        "discount": deal.discount,
        "discount_type": deal.discount_type or "percent",
        "tax_rate": deal.tax_rate,
        "tax_included": deal.tax_included,
        "deal_date": deal.deal_date.isoformat() if deal.deal_date else None,
        "address": deal.address or "",
        "repeat_interval_days": deal.repeat_interval_days,
        "next_repeat_date": deal.next_repeat_date.isoformat() if deal.next_repeat_date else None,
    }

@app.patch("/api/deals/{deal_id}")
def update_deal(deal_id: int, deal_data: DealUpdate, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    deal = db.query(Deal).filter(Deal.id == deal_id).first()
    if not deal: raise HTTPException(404, "Сделка не найдена")

    update_data = deal_data.model_dump(exclude_unset=True)

    if "new_contact_name" in update_data:
        new_contact = Contact(name=update_data["new_contact_name"])
        db.add(new_contact)
        db.flush()
        db.refresh(new_contact)
        deal.contact_id = new_contact.id
    elif "contact_id" in update_data:
        deal.contact_id = update_data["contact_id"]

    recalculate = "services" in update_data or "discount" in update_data or "discount_type" in update_data or "tax_rate" in update_data or "tax_included" in update_data

    if "services" in update_data:
        db.query(DealService).filter(DealService.deal_id == deal_id).delete(synchronize_session=False)
        for item_data in update_data["services"]:
            item = DealServiceItem(**item_data)
            service = db.query(Service).filter(Service.id == item.service_id).first()
            if service:
                db.add(DealService(deal_id=deal_id, service_id=service.id, quantity=item.quantity, price_at_moment=(service.price or 0)))
        db.flush()
    
    if "discount" in update_data: deal.discount = update_data["discount"]
    if "discount_type" in update_data: deal.discount_type = update_data["discount_type"]
    if "tax_rate" in update_data: deal.tax_rate = update_data["tax_rate"]
    if "tax_included" in update_data: deal.tax_included = update_data["tax_included"]
    
    if recalculate:
        subtotal = sum(ds.price_at_moment * ds.quantity for ds in deal.services)
        discount_val = deal.discount or 0
        discount_type = deal.discount_type or "percent"
        tax_rate_percent = deal.tax_rate or 0
        if discount_type == "fixed":
            discount_amount = min(discount_val, subtotal)
        else:
            discount_amount = subtotal * (discount_val / 100.0)
        subtotal_after_discount = subtotal - discount_amount
        
        tax_amount = subtotal_after_discount * (tax_rate_percent / 100.0)
        if deal.tax_included:
            final_total = subtotal_after_discount
        else:
            final_total = subtotal_after_discount + tax_amount
        
        deal.total = round(final_total, 2)

    if "title" in update_data: deal.title = update_data["title"]
    if "stage_id" in update_data: deal.stage_id = update_data["stage_id"]
    if "manager" in update_data: deal.manager = update_data["manager"]
    if "address" in update_data: deal.address = update_data["address"]
    if "work_date" in update_data or "work_time" in update_data:
        work_date = update_data.get("work_date") or (deal.deal_date.date().isoformat() if deal.deal_date else None)
        work_time = update_data.get("work_time") or (deal.deal_date.strftime("%H:%M") if deal.deal_date else None)
        if work_date:
            try:
                dt_str = work_date
                if work_time:
                    dt_str += f"T{work_time}"
                deal.deal_date = datetime.fromisoformat(dt_str)
            except Exception:
                pass
    if "repeat_interval_days" in update_data:
        deal.repeat_interval_days = update_data["repeat_interval_days"] or None
    if "next_repeat_date" in update_data:
        raw = update_data["next_repeat_date"]
        if raw:
            try:
                from datetime import date as _d
                deal.next_repeat_date = _d.fromisoformat(str(raw)[:10])
            except Exception:
                pass
        else:
            deal.next_repeat_date = None

    # Если сделка перешла в финальный неуспешный этап — снять повтор
    if "stage_id" in update_data and update_data["stage_id"]:
        stage = db.query(Stage).filter(Stage.id == update_data["stage_id"]).first()
        if stage and stage.is_final and "успешно" not in (stage.name or "").lower():
            deal.repeat_interval_days = None
            deal.next_repeat_date = None

    db.commit()
    db.refresh(deal)
    _cache.invalidate("deals", "years")
    return {"status": "ok", "id": deal.id}

@app.delete("/api/deals/{deal_id}", status_code=204)
def delete_deal(deal_id: int, db:DBSession=Depends(get_db),_=Depends(require_admin)):
    deal=db.query(Deal).filter(Deal.id==deal_id).first()
    if deal:
        db.query(DealService).filter(DealService.deal_id==deal_id).delete(synchronize_session=False)
        db.delete(deal); db.commit(); _cache.invalidate("deals","years")
    return None

# --- CONTACTS ---
@app.get("/api/contacts")
def get_contacts(search: Optional[str] = None, db:DBSession=Depends(get_db),_=Depends(get_current_user)):
    import json as _json
    q = db.query(Contact).order_by(Contact.name)
    if search and search.strip():
        s = f"%{search.strip()}%"
        q = q.filter(
            Contact.name.ilike(s) |
            Contact.phone.ilike(s) |
            Contact.source.ilike(s)
        )
    contacts = q.all()
    result = []
    for c in contacts:
        d = {col.name: getattr(c, col.name) for col in c.__table__.columns}
        try: d['addresses'] = _json.loads(c.addresses) if c.addresses else []
        except: d['addresses'] = []
        result.append(d)
    return result

@app.post("/api/contacts", status_code=201)
def create_contact(contact_data: ContactCreate, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    import json as _json
    if contact_data.phone and contact_data.phone.strip():
        existing = db.query(Contact).filter(Contact.phone == contact_data.phone).first()
        if existing: raise HTTPException(status_code=409, detail="Контакт с таким телефоном уже существует")
    data = contact_data.model_dump()
    addresses = data.pop('addresses', None) or []
    new_contact = Contact(**data, addresses=_json.dumps(addresses, ensure_ascii=False))
    db.add(new_contact); db.commit(); db.refresh(new_contact)
    _cache.invalidate("contacts")
    result = {col.name: getattr(new_contact, col.name) for col in new_contact.__table__.columns}
    result['addresses'] = addresses
    return result

@app.patch("/api/contacts/{contact_id}")
def update_contact(contact_id: int, contact_data: ContactUpdate, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    import json as _json
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact: raise HTTPException(status_code=404, detail="Контакт не найден")
    
    update_data = contact_data.model_dump(exclude_unset=True)
    if "phone" in update_data and update_data["phone"] and update_data["phone"].strip():
        existing = db.query(Contact).filter(Contact.phone == update_data["phone"], Contact.id != contact_id).first()
        if existing: raise HTTPException(status_code=409, detail="Контакт с таким телефоном уже существует")

    if 'addresses' in update_data:
        contact.addresses = _json.dumps(update_data.pop('addresses') or [], ensure_ascii=False)

    for key, value in update_data.items(): setattr(contact, key, value)
    
    db.commit(); db.refresh(contact)
    _cache.invalidate("contacts", "deals")
    result = {col.name: getattr(contact, col.name) for col in contact.__table__.columns}
    try: result['addresses'] = _json.loads(contact.addresses) if contact.addresses else []
    except: result['addresses'] = []
    return result


@app.get("/api/contacts/{contact_id}/deals")
def get_contact_all_deals(contact_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    """Все сделки контакта за всё время, от новых к старым"""
    deals = (db.query(Deal)
             .options(joinedload(Deal.stage))
             .filter(Deal.contact_id == contact_id)
             .order_by(Deal.created_at.desc())
             .all())
    won_names = {s.name for s in db.query(Stage).all() if s.is_final and "успешно" in (s.name or "").lower()}
    result = []
    for d in deals:
        result.append({
            "id": d.id,
            "title": d.title or "",
            "total": d.total or 0.0,
            "stage": d.stage.name if d.stage else "",
            "stage_color": d.stage.color if d.stage else "#6b7280",
            "stage_is_won": (d.stage.name in won_names) if d.stage else False,
            "stage_is_final": d.stage.is_final if d.stage else False,
            "deal_date": d.deal_date.isoformat() if d.deal_date else None,
            "created_at": (d.created_at or datetime.utcnow()).isoformat(),
            "repeat_interval_days": d.repeat_interval_days,
            "next_repeat_date": d.next_repeat_date.isoformat() if d.next_repeat_date else None,
        })
    total_revenue = sum(r["total"] for r in result if r["stage_is_won"])
    return {
        "deals": result,
        "total_count": len(result),
        "won_count": sum(1 for r in result if r["stage_is_won"]),
        "total_revenue": round(total_revenue, 2),
    }

@app.delete("/api/contacts/{contact_id}", status_code=204)
def delete_contact(contact_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact: return None
    if db.query(Deal).filter(Deal.contact_id == contact_id).count() > 0:
        raise HTTPException(status_code=400, detail="Нельзя удалить контакт, к которому привязаны сделки.")
    db.delete(contact); db.commit()
    _cache.invalidate("contacts"); return None

# --- EXPENSES ---
@app.get("/api/expense-categories")
def get_expense_categories(db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    return db.query(ExpenseCategory).order_by(ExpenseCategory.name).all()

@app.post("/api/expenses", status_code=201)
def create_expense(data: ExpenseCreate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    category_name = data.category.strip()
    category = None
    if category_name:
        category = db.query(ExpenseCategory).filter(func.lower(ExpenseCategory.name) == func.lower(category_name)).first()
        if not category:
            category = ExpenseCategory(name=category_name)
            db.add(category)
            db.flush()
            _cache.invalidate("expense_categories")

    from datetime import date as _dt
    try:
        parsed_date = _dt.fromisoformat(data.date[:10]) if data.date and data.date not in ("null","None","") else _dt.today()
    except (ValueError, TypeError):
        parsed_date = _dt.today()
    new_expense = Expense(name=data.name, amount=data.amount, date=parsed_date, category_id=category.id if category else None)
    db.add(new_expense); db.commit(); db.refresh(new_expense)
    _cache.invalidate("expenses", "years")
    return {
        "id": new_expense.id,
        "name": new_expense.name,
        "amount": new_expense.amount,
        "date": new_expense.date.isoformat() if new_expense.date else None,
        "category": new_expense.category.name if new_expense.category else ""
    }

@app.patch("/api/expenses/{expense_id}")
def update_expense(expense_id: int, data: ExpenseUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense: raise HTTPException(404, "Расход не найден")
    
    update_data = data.model_dump(exclude_unset=True)
    if "category" in update_data:
        category_name = update_data.pop("category").strip()
        category = None
        if category_name:
            category = db.query(ExpenseCategory).filter(func.lower(ExpenseCategory.name) == func.lower(category_name)).first()
            if not category:
                category = ExpenseCategory(name=category_name)
                db.add(category); db.flush()
                _cache.invalidate("expense_categories")
        expense.category_id = category.id if category else None

    if "date" in update_data:
        raw_date = update_data.pop("date")
        if raw_date and raw_date not in ("null", "None", ""):
            try:
                from datetime import date as _date
                expense.date = _date.fromisoformat(raw_date[:10])
            except (ValueError, TypeError):
                pass  # keep existing date

    for key, value in update_data.items(): setattr(expense, key, value)
    
    db.commit(); db.refresh(expense)
    _cache.invalidate("expenses", "years")
    return {
        "id": expense.id,
        "name": expense.name,
        "amount": expense.amount,
        "date": expense.date.isoformat() if expense.date else None,
        "category": expense.category.name if expense.category else ""
    }

@app.delete("/api/expenses/{expense_id}", status_code=204)
def delete_expense(expense_id: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if expense: db.delete(expense); db.commit(); _cache.invalidate("expenses", "years")
    return None

@app.get("/api/expenses")
def get_expenses(year: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(Expense).options(joinedload(Expense.category)).filter(extract("year", Expense.date) == year).order_by(Expense.date.desc())
    return [{"id": e.id, "name": e.name, "amount": e.amount, "category": e.category.name if e.category else "", "date": e.date.isoformat() if e.date else None} for e in q.all()]


# --- EQUIPMENT & MAINTENANCE ---
@app.get("/api/equipment", response_model=List[EquipmentResponse])
def get_equipment(db: DBSession=Depends(get_db), _=Depends(get_current_user)):
    return db.query(Equipment).order_by(Equipment.name).all()

@app.post("/api/equipment", status_code=201, response_model=EquipmentResponse)
def create_equipment(data: EquipmentCreate, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    new_item = Equipment(**data.model_dump()); db.add(new_item); db.commit(); db.refresh(new_item)
    _cache.invalidate("equipment"); return new_item

@app.patch("/api/equipment/{eq_id}", response_model=EquipmentResponse)
def update_equipment(eq_id: int, data: EquipmentUpdate, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    item = db.query(Equipment).filter(Equipment.id == eq_id).first()
    if not item: raise HTTPException(404, "Техника не найдена")
    for key, value in data.model_dump(exclude_unset=True).items(): setattr(item, key, value)
    db.commit(); db.refresh(item)
    _cache.invalidate("equipment"); return item

@app.delete("/api/equipment/{eq_id}", status_code=204)
def delete_equipment(eq_id: int, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    item = db.query(Equipment).filter(Equipment.id == eq_id).first()
    if item: db.delete(item); db.commit(); _cache.invalidate("equipment")
    return None

@app.get("/api/maintenance", response_model=List[MaintenanceForListResponse])
def get_all_maintenance(year: Optional[int] = None, db: DBSession=Depends(get_db), _=Depends(get_current_user)):
    q = db.query(EquipmentMaintenance).options(joinedload(EquipmentMaintenance.equipment)).order_by(EquipmentMaintenance.date.desc())
    if year: q = q.filter(extract("year", EquipmentMaintenance.date) == year)
    return q.all()

@app.get("/api/maintenance/{m_id}", response_model=MaintenanceDetailResponse)
def get_maintenance_details(m_id: int, db:DBSession=Depends(get_db), _=Depends(get_current_user)):
    m_record = db.query(EquipmentMaintenance).options(joinedload(EquipmentMaintenance.consumables_used).joinedload(MaintenanceConsumable.consumable), joinedload(EquipmentMaintenance.equipment)).filter(EquipmentMaintenance.id == m_id).first()
    if not m_record: raise HTTPException(404, "Запись о ТО не найдена")
    return m_record

@app.post("/api/maintenance", status_code=201, response_model=MaintenanceDetailResponse)
def create_maintenance_record(data: MaintenanceCreate, db:DBSession=Depends(get_db), _=Depends(get_current_user)):
    total_cost = 0
    try:
        for item_data in data.consumables:
            consumable = db.query(Consumable).filter(Consumable.id == item_data.consumable_id).with_for_update().first()
            if not consumable or consumable.stock_quantity < item_data.quantity:
                raise HTTPException(400, f"Недостаточно '{consumable.name if consumable else 'ID:'+str(item_data.consumable_id)}' на складе.")
            consumable.stock_quantity -= item_data.quantity
            total_cost += (consumable.price or 0) * item_data.quantity
        
        new_item = EquipmentMaintenance(equipment_id=data.equipment_id, date=data.date, work_description=data.work_description, notes=data.notes, cost=total_cost)
        db.add(new_item)
        db.flush()

        for item_data in data.consumables:
            consumable = db.query(Consumable).filter(Consumable.id == item_data.consumable_id).first()
            db.add(MaintenanceConsumable(maintenance_id=new_item.id, consumable_id=item_data.consumable_id, quantity=item_data.quantity, price_at_moment=(consumable.price or 0)))
        
        db.commit()
        db.refresh(new_item)
        update_equipment_last_maintenance(db, data.equipment_id)
        _cache.invalidate("consumables", "equipment", "maintenance")
        return new_item
    except:
        db.rollback()
        raise

@app.patch("/api/maintenance/{m_id}", response_model=MaintenanceDetailResponse)
def update_maintenance_record(m_id: int, data: MaintenanceUpdate, db:DBSession=Depends(get_db), _=Depends(get_current_user)):
    try:
        m_record = db.query(EquipmentMaintenance).options(joinedload(EquipmentMaintenance.consumables_used)).filter(EquipmentMaintenance.id == m_id).first()
        if not m_record: raise HTTPException(404, "Запись о ТО не найдена")

        if data.consumables is not None:
            for old_item in m_record.consumables_used:
                consumable = db.query(Consumable).filter(Consumable.id == old_item.consumable_id).with_for_update().first()
                if consumable: consumable.stock_quantity += old_item.quantity
            
            db.query(MaintenanceConsumable).filter(MaintenanceConsumable.maintenance_id == m_id).delete(synchronize_session=False)
            db.flush()

            total_cost = 0
            for item_data in data.consumables:
                consumable = db.query(Consumable).filter(Consumable.id == item_data.consumable_id).with_for_update().first()
                if not consumable or consumable.stock_quantity < item_data.quantity:
                    raise HTTPException(400, f"Недостаточно '{consumable.name if consumable else 'ID:'+str(item_data.consumable_id)}' на складе.")
                consumable.stock_quantity -= item_data.quantity
                total_cost += (consumable.price or 0) * item_data.quantity
                db.add(MaintenanceConsumable(maintenance_id=m_id, consumable_id=item_data.consumable_id, quantity=item_data.quantity, price_at_moment=(consumable.price or 0)))
            m_record.cost = total_cost

        update_data = data.model_dump(exclude_unset=True)
        if 'date' in update_data: m_record.date = update_data['date']
        if 'work_description' in update_data: m_record.work_description = update_data['work_description']
        if 'notes' in update_data: m_record.notes = update_data['notes']
        
        db.commit()
        db.refresh(m_record)
        update_equipment_last_maintenance(db, m_record.equipment_id)
        _cache.invalidate("consumables", "equipment", "maintenance")
        return m_record
    except:
        db.rollback()
        raise

@app.delete("/api/maintenance/{m_id}", status_code=204)
def delete_maintenance_record(m_id: int, db:DBSession=Depends(get_db), _=Depends(get_current_user)):
    try:
        item = db.query(EquipmentMaintenance).options(joinedload(EquipmentMaintenance.consumables_used)).filter(EquipmentMaintenance.id == m_id).first()
        if item:
            equipment_id = item.equipment_id
            for used in item.consumables_used:
                consumable = db.query(Consumable).filter(Consumable.id == used.consumable_id).with_for_update().first()
                if consumable: consumable.stock_quantity += used.quantity
            
            db.delete(item)
            db.commit()
            update_equipment_last_maintenance(db, equipment_id)
            _cache.invalidate("consumables", "equipment", "maintenance")
    except:
        db.rollback()
        raise
    return None

# --- CONSUMABLES ---
@app.get("/api/consumables")
def get_consumables(db:DBSession=Depends(get_db), _=Depends(get_current_user)):
    return db.query(Consumable).order_by(Consumable.name).all()

@app.post("/api/consumables", status_code=201)
def create_consumable(data: ConsumableCreate, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    new_item = Consumable(**data.model_dump()); db.add(new_item); db.commit(); db.refresh(new_item)
    _cache.invalidate("consumables"); return new_item

@app.patch("/api/consumables/{c_id}")
def update_consumable(c_id: int, data: ConsumableUpdate, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    item = db.query(Consumable).filter(Consumable.id == c_id).first()
    if not item: raise HTTPException(404, "Расходник не найден")
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items(): setattr(item, key, value)
    db.commit(); db.refresh(item)
    _cache.invalidate("consumables"); return item

@app.delete("/api/consumables/{c_id}", status_code=204)
def delete_consumable(c_id: int, db:DBSession=Depends(get_db), _=Depends(require_admin)):
    if db.query(MaintenanceConsumable).filter(MaintenanceConsumable.consumable_id == c_id).count() > 0:
        raise HTTPException(400, "Нельзя удалить расходник, который используется в записях о ТО.")
    item = db.query(Consumable).filter(Consumable.id == c_id).first()
    if item: db.delete(item); db.commit(); _cache.invalidate("consumables")

# --- OTHER ---
@app.get("/api/years")
def get_years(db: DBSession=Depends(get_db),_=Depends(get_current_user)):
    if (cached := _cache.get("years")) is not None: return cached
    years=sorted(set([r[0] for r in db.execute(text("SELECT DISTINCT EXTRACT(YEAR FROM deal_date)::int FROM deals WHERE deal_date IS NOT NULL")).fetchall() if r[0]] + [r[0] for r in db.execute(text("SELECT DISTINCT EXTRACT(YEAR FROM date)::int FROM expenses WHERE date IS NOT NULL")).fetchall() if r[0]]), reverse=True) or [datetime.utcnow().year]
    _cache.set("years", years); return years

@app.get("/api/tasks")
def get_tasks(year: Optional[int] = None, is_done: Optional[bool] = None, priority: Optional[str] = None, contact_id: Optional[int] = None, deal_id: Optional[int] = None, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    q = db.query(Task).options(joinedload(Task.contact), joinedload(Task.deal)).order_by(Task.due_date.asc())
    if year: q = q.filter(extract("year", Task.due_date) == year)
    if is_done is not None: q = q.filter(Task.is_done == is_done)
    if priority: q = q.filter(Task.priority == priority)
    if contact_id: q = q.filter(Task.contact_id == contact_id)
    if deal_id: q = q.filter(Task.deal_id == deal_id)
    if not is_admin(user): q = q.filter(Task.assignee == user["sub"])
    users = {u.id: u.name for u in db.query(User).all()}
    today = date.today()
    tasks_with_names = []
    for t in q.all():
        task_dict = {c.name: getattr(t, c.name) for c in t.__table__.columns}
        task_dict["assignee_name"] = users.get(t.assignee, "Не назначен")
        task_dict["contact_name"] = t.contact.name if t.contact else None
        task_dict["deal_title"] = t.deal.title if t.deal else None
        task_dict["is_overdue"] = bool(t.due_date and t.due_date < today and not t.is_done)
        tasks_with_names.append(task_dict)
    return tasks_with_names

@app.post("/api/tasks", status_code=201)
def create_task(task: TaskCreate, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    new_task = Task(
        title=task.title, description=task.description,
        due_date=task.due_date or (date.today() + timedelta(days=1)),
        priority=task.priority, status=task.status,
        assignee=task.assignee or user['sub'],
        contact_id=task.contact_id, deal_id=task.deal_id
    )
    db.add(new_task); db.commit(); db.refresh(new_task)
    _cache.invalidate("tasks"); return new_task

@app.patch("/api/tasks/{task_id}")
def update_task(task_id: int, task_data: TaskUpdate, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task: raise HTTPException(404, "Task not found")
    for key, value in task_data.model_dump(exclude_unset=True).items(): setattr(task, key, value)
    if task_data.is_done: task.status = "Выполнена"
    elif task_data.status == "Выполнена": task.is_done = True
    db.commit(); _cache.invalidate("tasks"); return {"status": "ok"}

@app.delete("/api/tasks/{task_id}", status_code=204)
def delete_task(task_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if task: db.delete(task); db.commit(); _cache.invalidate("tasks")

# --- TAXES ---
@app.get("/api/taxes/summary")
def get_tax_summary(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    cache_key = f"tax_summary:{year}"
    if (cached := _cache.get(cache_key)) is not None: return cached

    success_stage = db.query(Stage).filter(Stage.name == "Успешно").first()
    if not success_stage: raise HTTPException(status_code=404, detail="Этап 'Успешно' не найден для расчета налога.")

    total_revenue = db.query(func.sum(Deal.total)).filter(
        Deal.stage_id == success_stage.id,
        extract("year", Deal.deal_date) == year
    ).scalar() or 0

    tax_accrued = total_revenue * TAX_RATE

    total_paid = db.query(func.sum(TaxPayment.amount)).filter(TaxPayment.year == year).scalar() or 0

    summary = {
        "revenue": round(total_revenue, 2),
        "tax_accrued": round(tax_accrued, 2),
        "paid": round(total_paid, 2),
        "balance": round(tax_accrued - total_paid, 2)
    }
    _cache.set(cache_key, summary)
    return summary

@app.get("/api/taxes/payments")
def get_tax_payments(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    cache_key = f"tax_payments:{year}"
    if (cached := _cache.get(cache_key)) is not None: return cached
    
    payments = db.query(TaxPayment).filter(TaxPayment.year == year).order_by(TaxPayment.date.desc()).all()
    _cache.set(cache_key, payments)
    return payments

@app.post("/api/taxes/payments", status_code=201)
def create_tax_payment(payment_data: TaxPaymentCreate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    if payment_data.amount <= 0: raise HTTPException(status_code=400, detail="Сумма платежа должна быть положительной.")
    new_payment = TaxPayment(**payment_data.model_dump())
    db.add(new_payment); db.commit(); db.refresh(new_payment)
    _cache.invalidate(f"tax_summary:{payment_data.year}", f"tax_payments:{payment_data.year}", "years")
    return new_payment

class TaxPaymentUpdate(BaseModel):
    amount: Optional[float] = None
    date: Optional[date] = None
    note: Optional[str] = None

@app.patch("/api/taxes/payments/{payment_id}")
def update_tax_payment(payment_id: int, data: TaxPaymentUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    payment = db.query(TaxPayment).filter(TaxPayment.id == payment_id).first()
    if not payment: raise HTTPException(404, "Платёж не найден")
    if data.amount is not None and data.amount <= 0: raise HTTPException(400, "Сумма должна быть положительной.")
    for key, value in data.model_dump(exclude_unset=True).items(): setattr(payment, key, value)
    db.commit(); db.refresh(payment)
    _cache.invalidate(f"tax_summary:{payment.year}", f"tax_payments:{payment.year}")
    return payment

@app.delete("/api/taxes/payments/{payment_id}", status_code=204)
def delete_tax_payment(payment_id: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    payment = db.query(TaxPayment).filter(TaxPayment.id == payment_id).first()
    if payment:
        year = payment.year
        db.delete(payment); db.commit()
        _cache.invalidate(f"tax_summary:{year}", f"tax_payments:{year}")
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# НОВЫЕ ФИЧИ v13.0: Аналитика, Экспорт, Бюджет
# ═══════════════════════════════════════════════════════════════════════════════

# ── МОДЕЛЬ БЮДЖЕТ ─────────────────────────────────────────────────────────────
class Budget(Base):
    __tablename__ = "budgets"
    id = Column(Integer, primary_key=True)
    year = Column(Integer, nullable=False)
    period = Column(String(20), nullable=False)   # "year" | "q1"-"q4" | "jan"-"dec"
    name = Column(String(200), nullable=False)
    planned_revenue = Column(Float, default=0.0)
    planned_expenses = Column(Float, default=0.0)
    notes = Column(Text)

# Создать новую таблицу если ещё нет
from sqlalchemy import inspect as sa_inspect
def _ensure_budget_table():
    insp = sa_inspect(engine)
    if not insp.has_table("budgets"):
        Budget.__table__.create(engine)


def _ensure_users_telegram_column():
    insp = sa_inspect(engine)
    if not insp.has_table("users"):
        return
    cols = {c["name"] for c in insp.get_columns("users")}
    with engine.begin() as conn:
        if "telegram_id" not in cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN telegram_id VARCHAR(50)"))
        try:
            conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_users_telegram_id ON users (telegram_id)"))
        except Exception:
            pass


def _ensure_user_last_login():
    with SessionFactory() as db:
        _add_column_if_missing(db, "users", "last_login", "TIMESTAMP")


def _add_column_if_missing(db, table: str, column: str, col_type: str):
    """Check information_schema before ALTER TABLE to avoid errors."""
    try:
        result = db.execute(text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name=:t AND column_name=:c"
        ), {"t": table, "c": column}).fetchone()
        if not result:
            db.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
            db.commit()
            print(f"Added column {table}.{column}", flush=True)
    except Exception as e:
        db.rollback()
        print(f"_add_column_if_missing {table}.{column}: {e}", flush=True)


def _ensure_repeat_columns():
    """Add repeat_interval_days and next_repeat_date to deals if missing."""
    with SessionFactory() as db:
        _add_column_if_missing(db, "deals", "repeat_interval_days", "INTEGER")
        _add_column_if_missing(db, "deals", "next_repeat_date", "DATE")


def _send_tg_sync(text_msg: str):
    """Fire-and-forget sync telegram notification via bot отчётов."""
    import requests as _req
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat  = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat:
        return
    try:
        _req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat, "text": text_msg, "parse_mode": "HTML"},
            timeout=8
        )
    except Exception as e:
        print(f"TG notify error: {e}", flush=True)


def _repeat_deals_worker():
    """Background thread: every 60 min check and create repeat deals."""
    import time as _time
    while True:
        try:
            _process_repeat_deals()
        except Exception as e:
            print(f"repeat_deals_worker error: {e}", flush=True)
        _time.sleep(3600)


def _process_repeat_deals():
    from datetime import date as _date, timedelta as _td
    today = _date.today()
    horizon = today + _td(days=3)  # создаём за 3 дня до выезда

    with SessionFactory() as db:
        # Стадия "Провалена" — финальная, не успешная
        lost_stage_ids = {
            s.id for s in db.query(Stage).all()
            if s.is_final and "успешно" not in (s.name or "").lower()
        }
        first_stage = db.query(Stage).order_by(Stage.order).first()

        deals = (
            db.query(Deal)
            .filter(Deal.repeat_interval_days.isnot(None))
            .filter(Deal.next_repeat_date.isnot(None))
            .filter(Deal.next_repeat_date <= horizon)
            .all()
        )

        for deal in deals:
            # Прекратить если сделка провалена
            if deal.stage_id in lost_stage_ids:
                deal.repeat_interval_days = None
                deal.next_repeat_date = None
                db.commit()
                continue

            repeat_date = deal.next_repeat_date

            # Проверяем нет ли уже созданной сделки на эту дату (избегаем дублей)
            # Ищем по диапазону суток — deal_date может содержать любое время
            day_start = datetime.combine(repeat_date, datetime.min.time())
            day_end   = datetime.combine(repeat_date, datetime.max.time())
            existing = db.query(Deal).filter(
                Deal.contact_id == deal.contact_id,
                Deal.deal_date >= day_start,
                Deal.deal_date <= day_end,
                Deal.title == deal.title,
                Deal.id != deal.id,
                Deal.is_repeat == True,
            ).first()

            if not existing:
                from datetime import timedelta as _td2
                next_date = repeat_date + _td2(days=deal.repeat_interval_days)

                # Новая сделка получает триггер повтора, оригинал его теряет
                new_deal = Deal(
                    contact_id=deal.contact_id,
                    stage_id=first_stage.id if first_stage else deal.stage_id,
                    title=deal.title,
                    notes=deal.notes or "",
                    manager=deal.manager,
                    address=deal.address,
                    tax_rate=deal.tax_rate,
                    tax_included=deal.tax_included,
                    discount=deal.discount,
                    discount_type=deal.discount_type,
                    deal_date=datetime.combine(repeat_date, datetime.min.time()),
                    is_repeat=True,
                    repeat_interval_days=deal.repeat_interval_days,
                    next_repeat_date=next_date,
                )
                db.add(new_deal)
                db.flush()
                db.refresh(new_deal)

                # Копируем услуги
                for ds in deal.services:
                    db.add(DealService(
                        deal_id=new_deal.id,
                        service_id=ds.service_id,
                        quantity=ds.quantity,
                        price_at_moment=ds.price_at_moment
                    ))
                db.flush()

                # Пересчёт суммы
                subtotal = sum(s.price_at_moment * s.quantity for s in new_deal.services)
                disc = deal.discount or 0
                if deal.discount_type == "fixed":
                    disc_amt = min(disc, subtotal)
                else:
                    disc_amt = subtotal * (disc / 100.0)
                after_disc = subtotal - disc_amt
                tax_amt = after_disc * ((deal.tax_rate or 0) / 100.0)
                new_deal.total = round(after_disc + (0 if deal.tax_included else tax_amt), 2)

                # Убираем триггер из оригинальной сделки — он теперь в новой
                deal.repeat_interval_days = None
                deal.next_repeat_date = None

                db.commit()

                # Telegram уведомление
                contact = db.query(Contact).filter(Contact.id == deal.contact_id).first()
                _contact_name = contact.name if contact else "—"
                _phone = (contact.phone if contact else "") or ""
                _svc_lines = []
                for si in new_deal.services:
                    _qty = si.quantity or 1
                    _price = float(si.price_at_moment or 0)
                    _line = _price * _qty
                    _price_str = f" — {int(_line):,} \u00a0".replace(",", "\u00a0") if _price else ""
                    _svc_lines.append(f"  · {si.service.name} × {_qty}{_price_str}")
                _lines = [f"<b>Сделка №{new_deal.id} — {deal.title}</b> (повтор)"]
                _lines.append(repeat_date.strftime("%d.%m.%Y"))
                _lines.append(f"Клиент: {_contact_name}")
                if _phone:
                    _lines.append(f"Телефон: {_phone}")
                if deal.address:
                    _lines.append(f"Адрес: {deal.address}")
                if _svc_lines:
                    _lines.append("")
                    _lines.append("Услуги:")
                    _lines.extend(_svc_lines)
                if new_deal.total:
                    _lines.append("")
                    _lines.append(f"Итого: {int(new_deal.total):,} ₽".replace(",", "\u00a0"))
                _lines.append("")
                _lines.append(f"Следующий повтор: {next_date.strftime('%d.%m.%Y')}")
                msg = "\n".join(_lines)
                threading.Thread(target=_send_tg_sync, args=(msg,), daemon=True).start()


def _ensure_deals_discount_type():
    insp = sa_inspect(engine)
    if not insp.has_table("deals"):
        return
    cols = {c["name"] for c in insp.get_columns("deals")}
    if "discount_type" not in cols:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE deals ADD COLUMN discount_type VARCHAR(10) DEFAULT 'percent' NOT NULL"))

def _ensure_files_kind_column():
    insp = sa_inspect(engine)
    if not insp.has_table("crm_files"):
        return
    cols = {c["name"] for c in insp.get_columns("crm_files")}
    if "file_kind" in cols:
        return
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE crm_files ADD COLUMN file_kind VARCHAR(20)"))

def _ensure_contacts_telegram_columns():
    insp = sa_inspect(engine)
    if not insp.has_table("contacts"):
        return
    cols = {c["name"] for c in insp.get_columns("contacts")}
    with engine.begin() as conn:
        if "telegram_id" not in cols:
            conn.execute(text("ALTER TABLE contacts ADD COLUMN telegram_id VARCHAR(50)"))
        if "telegram_username" not in cols:
            conn.execute(text("ALTER TABLE contacts ADD COLUMN telegram_username VARCHAR(100)"))
        if "addresses" not in cols:
            conn.execute(text("ALTER TABLE contacts ADD COLUMN addresses TEXT"))
        try:
            conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_contacts_telegram_id ON contacts (telegram_id) WHERE telegram_id IS NOT NULL"))
        except Exception:
            pass


# ── PYDANTIC для бюджета ──────────────────────────────────────────────────────
class BudgetCreate(BaseModel):
    year: int
    period: str
    name: str
    planned_revenue: Optional[float] = 0.0
    planned_expenses: Optional[float] = 0.0
    notes: Optional[str] = None

class BudgetUpdate(BaseModel):
    name: Optional[str] = None
    planned_revenue: Optional[float] = None
    planned_expenses: Optional[float] = None
    notes: Optional[str] = None

# ── АНАЛИТИКА ─────────────────────────────────────────────────────────────────
@app.get("/api/analytics/funnel")  # kept for backwards compat
def get_funnel(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    return get_analytics(year, db, _)

@app.get("/api/analytics")
def get_analytics(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    from collections import defaultdict
    MONTHS_RU = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]

    stages = db.query(Stage).order_by(Stage.order).all()
    stage_map = {s.id: s for s in stages}

    deals = (db.query(Deal)
             .options(joinedload(Deal.services).joinedload(DealService.service),
                      joinedload(Deal.contact))
             .filter(extract("year", Deal.created_at) == year)
             .all())

    won_stage_ids  = {s.id for s in stages if s.is_final and "успешно" in s.name.lower()}
    lost_stage_ids = {s.id for s in stages if s.is_final and "успешно" not in s.name.lower()}
    won  = [d for d in deals if d.stage_id in won_stage_ids]
    lost = [d for d in deals if d.stage_id in lost_stage_ids]

    total_revenue = sum(d.total or 0 for d in won)
    avg_check     = round(total_revenue / len(won), 2) if won else 0
    win_rate      = round(len(won) / len(deals) * 100, 1) if deals else 0

    # 1. Воронка
    funnel = []
    for st in stages:
        sd = [d for d in deals if d.stage_id == st.id]
        funnel.append({
            "stage_id": st.id, "stage_name": st.name, "color": st.color,
            "is_final": st.is_final, "count": len(sd),
            "total": sum(d.total or 0 for d in sd),
        })

    # 2. Помесячная динамика
    rev_by_month = defaultdict(float)
    for d in won:
        m = (d.closed_at or d.created_at).month
        rev_by_month[m] += d.total or 0

    expenses_year = db.query(Expense).filter(extract("year", Expense.date) == year).all()
    exp_by_month  = defaultdict(float)
    for e in expenses_year:
        exp_by_month[e.date.month] += e.amount

    monthly = [
        {"month": i, "label": MONTHS_RU[i-1],
         "revenue":  round(rev_by_month[i], 2),
         "expenses": round(exp_by_month[i], 2),
         "profit":   round(rev_by_month[i] - exp_by_month[i], 2)}
        for i in range(1, 13)
    ]

    # 3. Топ услуг по выручке
    svc_revenue = defaultdict(float)
    svc_count   = defaultdict(int)
    svc_names   = {}
    for d in won:
        for ds in d.services:
            sname = ds.service.name if ds.service else f"#{ds.service_id}"
            svc_revenue[ds.service_id] += ds.price_at_moment * ds.quantity
            svc_count[ds.service_id]   += 1
            svc_names[ds.service_id]    = sname
    top_services = sorted(
        [{"id": sid, "name": svc_names[sid],
          "revenue": round(svc_revenue[sid], 2), "count": svc_count[sid]}
         for sid in svc_revenue],
        key=lambda x: x["revenue"], reverse=True
    )[:8]

    # 4. Расходы по категориям
    cats = {c.id: c.name for c in db.query(ExpenseCategory).all()}
    exp_by_cat = defaultdict(float)
    for e in expenses_year:
        exp_by_cat[cats.get(e.category_id, "Прочее")] += e.amount
    expense_by_category = sorted(
        [{"name": k, "amount": round(v, 2)} for k, v in exp_by_cat.items()],
        key=lambda x: x["amount"], reverse=True
    )

    # 5. Повторные клиенты
    repeat_won = [d for d in won if d.is_repeat]
    new_won    = [d for d in won if not d.is_repeat]
    repeat_rev = sum(d.total or 0 for d in repeat_won)
    new_rev    = sum(d.total or 0 for d in new_won)

    # 6. Причины провала и 7. Источники клиентов
    lost_reason_count = defaultdict(int)
    if lost_stage_ids:
        lost_comments = (
            db.query(DealComment.text)
            .join(Deal, Deal.id == DealComment.deal_id)
            .filter(
                extract("year", Deal.created_at) == year,
                Deal.stage_id.in_(lost_stage_ids),
                DealComment.text.ilike("Причина провала:%")
            )
            .all()
        )
        for (txt,) in lost_comments:
            text_val = (txt or "").strip()
            reason = text_val.split(":", 1)[1].strip() if ":" in text_val else text_val
            reason = reason or "Не указана"
            lost_reason_count[reason] += 1

    lost_reasons = sorted(
        [{"name": k, "count": v} for k, v in lost_reason_count.items()],
        key=lambda x: x["count"], reverse=True
    )

    client_source_count = defaultdict(int)
    unique_contact_ids = {d.contact_id for d in deals if d.contact_id}
    if unique_contact_ids:
        contacts = db.query(Contact).filter(Contact.id.in_(list(unique_contact_ids))).all()
        for c in contacts:
            src = (c.source or "").strip() or "Не указан"
            client_source_count[src] += 1

    client_sources = sorted(
        [{"name": k, "count": v} for k, v in client_source_count.items()],
        key=lambda x: x["count"], reverse=True
    )

    return {
        "total_deals":   len(deals),
        "won_deals":     len(won),
        "lost_deals":    len(lost),
        "win_rate":      win_rate,
        "avg_check":     avg_check,
        "total_revenue": round(total_revenue, 2),
        "total_expenses": round(sum(e.amount for e in expenses_year), 2),
        "funnel":              funnel,
        "monthly":             monthly,
        "top_services":        top_services,
        "expense_by_category": expense_by_category,
        "repeat": {
            "repeat_count":   len(repeat_won),
            "new_count":      len(new_won),
            "repeat_revenue": round(repeat_rev, 2),
            "new_revenue":    round(new_rev, 2),
            "repeat_rate":    round(len(repeat_won) / len(won) * 100, 1) if won else 0,
        },
        "lost_reasons": lost_reasons,
        "client_sources": client_sources,
    }


# ── БЮДЖЕТ — CRUD ─────────────────────────────────────────────────────────────
@app.get("/api/budget", response_model=List[BudgetResponse])
def get_budget(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    items = db.query(Budget).filter(Budget.year == year).order_by(Budget.id).all()
    return items

@app.post("/api/budget", status_code=201)
def create_budget(data: BudgetCreate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = Budget(**data.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item

@app.patch("/api/budget/{item_id}")
def update_budget(item_id: int, data: BudgetUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = db.query(Budget).filter(Budget.id == item_id).first()
    if not item: raise HTTPException(404, "Запись не найдена")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit(); db.refresh(item)
    return item

@app.delete("/api/budget/{item_id}", status_code=204)
def delete_budget(item_id: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = db.query(Budget).filter(Budget.id == item_id).first()
    if item: db.delete(item); db.commit()
    return None

# ── ЭКСПОРТ EXCEL ─────────────────────────────────────────────────────────────
from fastapi.responses import StreamingResponse
import io

def _openpyxl_available():
    try:
        import openpyxl
        return True
    except ImportError:
        return False

@app.get("/api/export/excel")
def export_excel(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    if not _openpyxl_available():
        raise HTTPException(503, "openpyxl не установлен на сервере")
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # убрать дефолтный лист

    GREEN = "1a3318"; SAGE = "4a7c3f"; LIGHT = "f5f2eb"; WHITE = "FFFFFF"
    
    def make_header(ws, cols, header_color=GREEN):
        for c, (title, width) in enumerate(cols, 1):
            cell = ws.cell(1, c, title)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=header_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.row_dimensions[1].height = 28

    def style_row(ws, row, r, alt=False):
        for c in range(1, len(row)+1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor="F0ECE3" if alt else WHITE)
            cell.alignment = Alignment(vertical="center")
            thin = Side(style="thin", color="E2DDD4")
            cell.border = Border(bottom=Border(bottom=thin).bottom)

    # ── Лист 1: Сделки ────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Сделки")
    stages = {s.id: s for s in db.query(Stage).all()}
    contacts = {c.id: c for c in db.query(Contact).all()}
    deals_q = db.query(Deal).filter(extract('year', Deal.created_at) == year).order_by(Deal.created_at).all()
    cols1 = [("ID",6),("Название",30),("Клиент",22),("Этап",18),("Сумма",14),("Дата создания",18),("Дата закрытия",18),("Менеджер",18)]
    make_header(ws1, cols1)
    for i, d in enumerate(deals_q, 2):
        row = [d.id, d.title, contacts.get(d.contact_id, type("x",(),{"name":"—"})()).name,
               stages.get(d.stage_id, type("x",(),{"name":"—"})()).name,
               d.total or 0,
               d.created_at.strftime("%d.%m.%Y") if d.created_at else "",
               d.closed_at.strftime("%d.%m.%Y") if d.closed_at else "",
               d.manager or ""]
        for c, v in enumerate(row, 1): ws1.cell(i, c, v)
        style_row(ws1, row, i, i%2==0)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:H{len(deals_q)+1}"

    # ── Лист 2: Расходы ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Расходы")
    expenses_q = db.query(Expense).filter(extract('year', Expense.date) == year).order_by(Expense.date).all()
    cats = {c.id: c.name for c in db.query(ExpenseCategory).all()}
    cols2 = [("ID",6),("Дата",14),("Название",35),("Категория",20),("Сумма",14)]
    make_header(ws2, cols2)
    for i, e in enumerate(expenses_q, 2):
        row = [e.id, e.date.strftime("%d.%m.%Y") if e.date else "", e.name, cats.get(e.category_id,"—"), e.amount]
        for c, v in enumerate(row, 1): ws2.cell(i, c, v)
        style_row(ws2, row, i, i%2==0)
    ws2.freeze_panes = "A2"
    
    # Итого расходов
    total_row = len(expenses_q)+2
    ws2.cell(total_row, 4, "ИТОГО").font = Font(bold=True)
    ws2.cell(total_row, 5, sum(e.amount for e in expenses_q)).font = Font(bold=True)

    # ── Лист 3: Аналитика ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Аналитика")
    won_stages = [s for s in stages.values() if s.is_final and "успешно" in s.name.lower()]
    won_ids = {s.id for s in won_stages}
    won_deals = [d for d in deals_q if d.stage_id in won_ids]
    total_rev = sum(d.total or 0 for d in won_deals)
    total_exp = sum(e.amount for e in expenses_q)
    avg_check = round(total_rev / len(won_deals), 2) if won_deals else 0

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 22
    summary_data = [
        ("Год", year), ("Всего сделок", len(deals_q)),
        ("Успешных сделок", len(won_deals)),
        ("Выручка", total_rev), ("Расходы", total_exp),
        ("Прибыль", total_rev - total_exp), ("Средний чек", avg_check),
    ]
    ws3.cell(1,1,"СВОДКА ЗА ГОД").font = Font(bold=True, size=13, color=GREEN)
    ws3.row_dimensions[1].height = 26
    for r, (label, val) in enumerate(summary_data, 2):
        c1 = ws3.cell(r, 1, label); c2 = ws3.cell(r, 2, val)
        c1.font = Font(bold=True, size=10)
        c2.fill = PatternFill("solid", fgColor=LIGHT)
        c2.alignment = Alignment(horizontal="right")

    ws3.cell(len(summary_data)+3, 1, "ВОРОНКА ПРОДАЖ").font = Font(bold=True, size=11, color=GREEN)
    funnel_cols = [("Этап",22),("Кол-во",12),("Сумма",16),("Конверсия от первого",22),("Конверсия от предыдущего",26)]
    frow = len(summary_data)+4
    for c,(t,w) in enumerate(funnel_cols,1):
        cell = ws3.cell(frow, c, t)
        cell.font = Font(bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=SAGE)
        ws3.column_dimensions[get_column_letter(c)].width = w
    
    stage_list = sorted(stages.values(), key=lambda s: s.order)
    first_count = None
    for i, st in enumerate(stage_list):
        cnt = sum(1 for d in deals_q if d.stage_id == st.id)
        amt = sum(d.total or 0 for d in deals_q if d.stage_id == st.id)
        if first_count is None: first_count = cnt
        conv_first = f"{round(cnt/first_count*100,1)}%" if first_count else "—"
        prev_cnt = sum(1 for d in deals_q if d.stage_id == stage_list[i-1].id) if i>0 else cnt
        conv_step = f"{round(cnt/prev_cnt*100,1)}%" if (i>0 and prev_cnt) else "100%"
        r = frow+1+i
        for c,v in enumerate([st.name, cnt, amt, conv_first, conv_step],1):
            ws3.cell(r, c, v)
        if (r-frow)%2==0:
            for c in range(1,6): ws3.cell(r,c).fill = PatternFill("solid", fgColor="F0ECE3")

    # ── Лист 4: Бюджет ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Бюджет")
    budgets_q = db.query(Budget).filter(Budget.year == year).all()
    cols4 = [("Название",28),("Период",14),("Плановая выручка",20),("Плановые расходы",20),("Плановая прибыль",20),("Заметки",30)]
    make_header(ws4, cols4)
    for i, b in enumerate(budgets_q, 2):
        profit = (b.planned_revenue or 0) - (b.planned_expenses or 0)
        row = [b.name, b.period, b.planned_revenue or 0, b.planned_expenses or 0, profit, b.notes or ""]
        for c, v in enumerate(row, 1): ws4.cell(i, c, v)
        style_row(ws4, row, i, i%2==0)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="grasscrm_{year}.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ── ЭКСПОРТ PDF ───────────────────────────────────────────────────────────────
@app.get("/api/export/pdf")
def export_pdf(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable, KeepTogether)
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        font_ok = False
        for fp in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                   "/usr/share/fonts/TTF/DejaVuSans.ttf"]:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("F", fp))
                except Exception:
                    pass  # already registered in this process - that's fine
                font_ok = True; break
        F = "F" if font_ok else "Helvetica"
    except ImportError:
        raise HTTPException(503, "reportlab не установлен на сервере")

    # ── данные ────────────────────────────────────────────────────────────────
    stages     = {s.id: s for s in db.query(Stage).all()}
    contacts   = {c.id: c for c in db.query(Contact).all()}
    deals_q    = db.query(Deal).options(joinedload(Deal.contact)).filter(
                     extract('year', Deal.created_at) == year).order_by(Deal.created_at).all()
    expenses_q = db.query(Expense).options(joinedload(Expense.category)).filter(
                     extract('year', Expense.date) == year).order_by(Expense.date).all()

    won_ids   = {s.id for s in stages.values() if s.is_final and "успешно" in s.name.lower()}
    lost_ids  = {s.id for s in stages.values() if s.is_final and "успешно" not in s.name.lower()}
    won_deals = [d for d in deals_q if d.stage_id in won_ids]
    lost_deals= [d for d in deals_q if d.stage_id in lost_ids]

    total_rev  = sum(d.total or 0 for d in won_deals)
    total_exp  = sum(e.amount or 0 for e in expenses_q)
    profit     = total_rev - total_exp
    avg_check  = total_rev / len(won_deals) if won_deals else 0
    win_rate   = round(len(won_deals) / len(deals_q) * 100, 1) if deals_q else 0
    MONTHS     = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]

    def rub(v): return f"{float(v or 0):,.0f}\u202f₽".replace(",", "\u202f")

    # ── цвета ─────────────────────────────────────────────────────────────────
    C_DARK   = colors.HexColor("#1a3318")
    C_GREEN  = colors.HexColor("#3d6b35")
    C_SAGE   = colors.HexColor("#4a7c3f")
    C_LIGHT  = colors.HexColor("#f5f2eb")
    C_ALT    = colors.HexColor("#ede9e0")
    C_WHITE  = colors.white
    C_MUTED  = colors.HexColor("#8a8070")
    C_BORDER = colors.HexColor("#d9d4c8")
    C_RED    = colors.HexColor("#c0392b")

    W = A4[0] - 4*cm   # полезная ширина страницы

    # ── стили параграфов ──────────────────────────────────────────────────────
    def ps(name, **kw):
        d = dict(fontName=F, fontSize=9, leading=13, textColor=C_DARK)
        d.update(kw); return ParagraphStyle(name, **d)

    S_LOGO   = ps("logo",  fontSize=20, textColor=C_DARK,  leading=24, spaceAfter=2)
    S_META   = ps("meta",  fontSize=8,  textColor=C_MUTED, spaceAfter=0)
    S_H2     = ps("h2",    fontSize=10, textColor=C_GREEN,  spaceBefore=18, spaceAfter=5,
                            borderPadding=(0,0,3,0))
    S_BODY   = ps("body",  fontSize=8.5, leading=12)
    S_FOOT   = ps("foot",  fontSize=7,  textColor=C_MUTED)
    S_KPI_L  = ps("kpil",  fontSize=7,  textColor=C_MUTED, leading=10, spaceAfter=1)
    S_KPI_V  = ps("kpiv",  fontSize=14, textColor=C_DARK,  leading=17, spaceAfter=0)
    S_KPI_V2 = ps("kpiv2", fontSize=11, textColor=C_SAGE,  leading=14, spaceAfter=0)

    # ── вспомогательные функции ───────────────────────────────────────────────
    def make_table(rows, widths, hdr=C_DARK, repeat=1):
        t = Table(rows, colWidths=widths, repeatRows=repeat)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  hdr),
            ("TEXTCOLOR",     (0,0), (-1,0),  C_WHITE),
            ("FONTNAME",      (0,0), (-1,-1), F),
            ("FONTSIZE",      (0,0), (-1,0),  7.5),
            ("FONTSIZE",      (0,1), (-1,-1), 8.5),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("RIGHTPADDING",  (0,0), (-1,-1), 8),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("LINEBELOW",     (0,0), (-1,0),  0.8, C_SAGE),
            ("LINEBELOW",     (0,1), (-1,-1), 0.3, C_BORDER),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_ALT]),
            ("BOX",           (0,0), (-1,-1), 0.5, C_BORDER),
        ]))
        return t

    # ── шапка страницы ────────────────────────────────────────────────────────
    story = []
    hdr_data = [[
        Paragraph("GrassCRM", S_LOGO),
        Paragraph(f"Отчёт за {year} год<br/><font color='#8a8070' size='8'>Сформирован {date.today().strftime('%d.%m.%Y')} · v2</font>", S_BODY)
    ]]
    hdr_t = Table(hdr_data, colWidths=[W*0.5, W*0.5])
    hdr_t.setStyle(TableStyle([
        ("FONTNAME",      (0,0), (-1,-1), F),
        ("VALIGN",        (0,0), (-1,-1), "BOTTOM"),
        ("ALIGN",         (1,0), (1,0),   "RIGHT"),
        ("TOPPADDING",    (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(hdr_t)
    story.append(HRFlowable(width=W, thickness=1.5, color=C_SAGE, spaceAfter=10, spaceBefore=6))

    # ── KPI-плитки ────────────────────────────────────────────────────────────
    def kpi(label, val, sub=None):
        items = [Paragraph(label, S_KPI_L), Paragraph(val, S_KPI_V)]
        if sub: items.append(Paragraph(sub, S_KPI_V2))
        return items

    kpi_rows = [
        kpi("ВЫРУЧКА",      rub(total_rev)),
        kpi("РАСХОДЫ",      rub(total_exp)),
        kpi("ПРИБЫЛЬ",      rub(profit)),
        kpi("СРЕДНИЙ ЧЕК",  rub(avg_check)),
    ]
    kpi_rows2 = [
        kpi("СДЕЛОК ВСЕГО",      str(len(deals_q))),
        kpi("КОНВЕРСИЯ",         f"{win_rate}%", f"успешных {len(won_deals)}"),
        kpi("ПРОВАЛЕНО",         str(len(lost_deals))),
        kpi("УНИКАЛЬНЫЕ КЛИЕНТЫ", str(len(set(d.contact_id for d in deals_q if d.contact_id)))),
    ]

    def kpi_table(data):
        rows = [[item[0] for item in data], [item[1] for item in data],
                [item[2] if len(item) > 2 else Paragraph("", S_KPI_V2) for item in data]]
        t = Table(rows, colWidths=[W/4]*4)
        t.setStyle(TableStyle([
            ("FONTNAME",      (0,0), (-1,-1), F),
            ("BACKGROUND",    (0,0), (-1,-1), C_LIGHT),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 12),
            ("RIGHTPADDING",  (0,0), (-1,-1), 12),
            ("LINEAFTER",     (0,0), (-2,-1), 0.5, C_BORDER),
            ("BOX",           (0,0), (-1,-1), 0.5, C_BORDER),
        ]))
        return t

    story.append(kpi_table(kpi_rows));  story.append(Spacer(1, 4))
    story.append(kpi_table(kpi_rows2)); story.append(Spacer(1, 2))

    # ── Помесячная динамика ───────────────────────────────────────────────────
    story.append(Paragraph("Помесячная динамика", S_H2))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    mrows = [["МЕСЯЦ", "ВЫРУЧКА", "РАСХОДЫ", "ПРИБЫЛЬ", "МАРЖА"]]
    for i, mn in enumerate(MONTHS, 1):
        mr = sum((d.total or 0) for d in won_deals if (d.closed_at or d.created_at).month == i)
        me = sum((e.amount or 0) for e in expenses_q if e.date and e.date.month == i)
        mg = f"{round((mr-me)/mr*100,1)}%" if mr else "—"
        mrows.append([mn, rub(mr), rub(me), rub(mr-me), mg])
    story.append(KeepTogether(make_table(mrows, [2.5*cm, 4.1*cm, 4.1*cm, 4.1*cm, 3.2*cm])))
    story.append(Spacer(1, 4))

    # ── Воронка ───────────────────────────────────────────────────────────────
    story.append(Paragraph("Воронка продаж", S_H2))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    frows = [["ЭТАП", "СДЕЛОК", "СУММА", "ДОЛЯ"]]
    total_cnt = len(deals_q) or 1
    for st in sorted(stages.values(), key=lambda s: s.order):
        cnt = sum(1 for d in deals_q if d.stage_id == st.id)
        amt = sum(d.total or 0 for d in deals_q if d.stage_id == st.id)
        frows.append([st.name, str(cnt), rub(amt), f"{round(cnt/total_cnt*100,1)}%"])
    story.append(KeepTogether(make_table(frows, [8*cm, 2.8*cm, 5*cm, 2.2*cm], C_SAGE)))
    story.append(Spacer(1, 4))

    # ── Топ-10 сделок ─────────────────────────────────────────────────────────
    story.append(Paragraph("Топ-10 успешных сделок", S_H2))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    drows = [["НАЗВАНИЕ", "КЛИЕНТ", "СУММА", "ЗАКРЫТА"]]
    for d in sorted(won_deals, key=lambda x: x.total or 0, reverse=True)[:10]:
        cname = contacts.get(d.contact_id, type("_",(),{"name":"—"})()).name
        closed = d.closed_at.strftime("%d.%m.%y") if d.closed_at else "—"
        drows.append([d.title[:38], cname[:22], rub(d.total or 0), closed])
    if len(drows) == 1: drows.append(["Нет данных", "—", "—", "—"])
    story.append(KeepTogether(make_table(drows, [8*cm, 4.5*cm, 3.5*cm, 2*cm])))
    story.append(Spacer(1, 4))

    # ── Расходы по категориям ─────────────────────────────────────────────────
    story.append(Paragraph("Расходы по категориям", S_H2))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    exp_by_cat: dict = {}
    for e in expenses_q:
        cat = e.category.name if e.category else "Без категории"
        exp_by_cat[cat] = exp_by_cat.get(cat, 0) + (e.amount or 0)
    crows = [["КАТЕГОРИЯ", "СУММА", "ДОЛЯ"]]
    te = total_exp or 1
    for cat, amt in sorted(exp_by_cat.items(), key=lambda x: x[1], reverse=True):
        crows.append([cat, rub(amt), f"{round(amt/te*100,1)}%"])
    if len(crows) == 1: crows.append(["Нет данных", rub(0), "0%"])
    story.append(KeepTogether(make_table(crows, [9.5*cm, 5*cm, 3.5*cm], C_SAGE)))

    # ── подвал ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    story.append(Paragraph(f"GrassCRM · crmpokos.ru · Отчёт {year} года", S_FOOT))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    doc.build(story)
    buf.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="grasscrm_{year}.pdf"',
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


# ── INTERACTIONS ───────────────────────────────────────────────────────────────

@app.get("/api/contacts/{contact_id}/interactions")
def get_interactions(contact_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(Interaction).filter(
        Interaction.contact_id == contact_id
    ).order_by(Interaction.created_at.desc()).all()
    return [
        {
            "id": i.id,
            "type": i.type,
            "text": i.text,
            "created_at": i.created_at.isoformat() if i.created_at else None,
            "user_name": i.user_name or "—"
        }
        for i in items
    ]

@app.post("/api/contacts/{contact_id}/interactions", status_code=201)
def create_interaction(
    contact_id: int,
    data: InteractionCreate,
    db: DBSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(404, "Контакт не найден")
    item = Interaction(
        contact_id=contact_id,
        type=data.type,
        text=data.text,
        user_id=user.get("sub"),
        user_name=user.get("name")
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {
        "id": item.id,
        "type": item.type,
        "text": item.text,
        "created_at": item.created_at.isoformat(),
        "user_name": item.user_name or "—"
    }

@app.delete("/api/interactions/{interaction_id}", status_code=204)
def delete_interaction(
    interaction_id: int,
    db: DBSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    item = db.query(Interaction).filter(Interaction.id == interaction_id).first()
    if not item:
        raise HTTPException(404, "Запись не найдена")
    if item.user_id != user.get("sub") and not is_admin(user):
        raise HTTPException(403, "Нет доступа")
    db.delete(item)
    db.commit()
    return None


# ── DEAL COMMENTS ──────────────────────────────────────────────────────────────

@app.get("/api/deals/{deal_id}/comments")
def get_deal_comments(deal_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(DealComment).filter(
        DealComment.deal_id == deal_id
    ).order_by(DealComment.created_at.asc()).all()
    return [
        {
            "id": i.id,
            "text": i.text,
            "created_at": i.created_at.isoformat() if i.created_at else None,
            "user_name": i.user_name or "—",
            "user_id": i.user_id
        }
        for i in items
    ]

@app.post("/api/deals/{deal_id}/comments", status_code=201)
def create_deal_comment(
    deal_id: int,
    data: DealCommentCreate,
    db: DBSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    deal = db.query(Deal).filter(Deal.id == deal_id).first()
    if not deal:
        raise HTTPException(404, "Сделка не найдена")
    item = DealComment(
        deal_id=deal_id,
        text=data.text,
        user_id=user.get("sub"),
        user_name=user.get("name")
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {
        "id": item.id,
        "text": item.text,
        "created_at": item.created_at.isoformat(),
        "user_name": item.user_name or "—",
        "user_id": item.user_id
    }

@app.delete("/api/deal-comments/{comment_id}", status_code=204)
def delete_deal_comment(
    comment_id: int,
    db: DBSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    item = db.query(DealComment).filter(DealComment.id == comment_id).first()
    if not item:
        raise HTTPException(404, "Комментарий не найден")
    if item.user_id != user.get("sub") and not is_admin(user):
        raise HTTPException(403, "Нет доступа")
    db.delete(item)
    db.commit()
    return None


# ── FILES ──────────────────────────────────────────────────────────────────────

ALLOWED_EXTENSIONS = {
    '.pdf','.doc','.docx','.xls','.xlsx','.ppt','.pptx',
    '.jpg','.jpeg','.png','.gif','.webp',
    '.txt','.csv','.zip','.rar',
    '.mp4','.mov','.avi'
}
HAS_MULTIPART = importlib.util.find_spec("multipart") is not None or importlib.util.find_spec("python_multipart") is not None

def _fmt_size(b: int) -> str:
    if b is None:
        return "—"
    if b < 1024: return f"{b} Б"
    if b < 1024**2: return f"{b/1024:.1f} КБ"
    return f"{b/1024**2:.1f} МБ"

@app.get("/api/files")
def get_files(contact_id: Optional[int] = None, deal_id: Optional[int] = None,
              db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(CRMFile).options(joinedload(CRMFile.contact), joinedload(CRMFile.deal)).order_by(CRMFile.created_at.desc())
    if contact_id: q = q.filter(CRMFile.contact_id == contact_id)
    if deal_id:    q = q.filter(CRMFile.deal_id == deal_id)
    return [{
        "id": f.id, "filename": f.filename, "size": f.size or 0,
        "size_fmt": _fmt_size(f.size), "mime_type": f.mime_type,
        "contact_id": f.contact_id, "deal_id": f.deal_id,
        "uploaded_by_name": f.uploaded_by_name or "—",
        "file_kind": f.file_kind or "general",
        "created_at": f.created_at.isoformat() if f.created_at else None,
        "contact_name": f.contact.name if f.contact else None,
        "deal_title": f.deal.title if f.deal else None,
    } for f in q.all()]

if HAS_MULTIPART:
    @app.post("/api/files", status_code=201)
    async def upload_file(
        file: UploadFile = FastAPIFile(...),
        contact_id: Optional[int] = Form(None),
        deal_id: Optional[int] = Form(None),
        file_kind: Optional[str] = Form(None),
        db: DBSession = Depends(get_db),
        user: dict = Depends(get_current_user)
    ):
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(400, f"Тип файла не разрешён: {ext}")

        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(400, "Файл слишком большой (макс. 20 МБ)")

        stored_name = f"{secrets.token_hex(12)}{ext}"
        dest = UPLOAD_DIR / stored_name
        dest.write_bytes(content)

        kind = (file_kind or "").strip().lower() or "general"
        if kind not in ("before", "after", "general"):
            kind = "general"

        rec = CRMFile(
            filename=file.filename,
            stored_name=stored_name,
            size=len(content),
            mime_type=file.content_type,
            contact_id=contact_id,
            deal_id=deal_id,
            uploaded_by=user.get("sub"),
            uploaded_by_name=user.get("name"),
            file_kind=kind,
        )
        db.add(rec); db.commit(); db.refresh(rec)
        return {
            "id": rec.id, "filename": rec.filename, "size": rec.size,
            "size_fmt": _fmt_size(rec.size), "mime_type": rec.mime_type,
        }
else:
    @app.post("/api/files", status_code=503)
    async def upload_file_unavailable(_=Depends(get_current_user)):
        raise HTTPException(503, "Загрузка файлов временно недоступна: установите python-multipart")

@app.get("/api/files/{file_id}/download")
def download_file(file_id: int, db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    from fastapi.responses import FileResponse as FR
    rec = db.query(CRMFile).filter(CRMFile.id == file_id).first()
    if not rec: raise HTTPException(404, "Файл не найден")
    path = UPLOAD_DIR / rec.stored_name
    if not path.exists(): raise HTTPException(404, "Файл удалён с диска")
    return FR(str(path), filename=rec.filename, media_type=rec.mime_type or "application/octet-stream")

@app.delete("/api/files/{file_id}", status_code=204)
def delete_file(file_id: int, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    rec = db.query(CRMFile).filter(CRMFile.id == file_id).first()
    if not rec: raise HTTPException(404, "Файл не найден")
    if rec.uploaded_by != user.get("sub") and not is_admin(user):
        raise HTTPException(403, "Нет доступа")
    path = UPLOAD_DIR / rec.stored_name
    if path.exists(): path.unlink()
    db.delete(rec); db.commit()
    return None


# ── SERVICE PANEL ─────────────────────────────────────────────────────────────

@app.get("/api/service/status")
def service_status(db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    
    system_payload = {}
    try:
        try:
            import psutil
        except ImportError:
            import sys as _sys
            for _p in ['/usr/lib/python3/dist-packages', '/usr/local/lib/python3.12/dist-packages',
                       '/usr/local/lib/python3.11/dist-packages', '/usr/local/lib/python3.10/dist-packages']:
                if _p not in _sys.path:
                    _sys.path.insert(0, _p)
            import psutil
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage("/")
        cpu = psutil.cpu_percent(interval=1.0)
        system_payload = {
            "cpu": round(cpu, 1),
            "mem_used_mb": round(mem.used / 1024**2),
            "mem_total_mb": round(mem.total / 1024**2),
            "mem_pct": round(mem.percent, 1),
            "disk_used_gb": round(disk.used / 1024**3, 1),
            "disk_total_gb": round(disk.total / 1024**3, 1),
            "disk_pct": round(disk.percent, 1),
        }
    except ImportError:
        system_payload = {"error": "psutil is not installed"}
    except Exception as e:
        system_payload = {"error": str(e)}

    active_tasks = db.query(Task).filter(Task.is_done == False).count()
    overdue = db.query(Task).filter(
        Task.is_done == False, Task.due_date < datetime.utcnow().date()
    ).count()
    active_deals = db.query(Deal).join(Stage).filter(Stage.is_final == False).count()
    cache_keys = list(_cache._data.keys())

    # Дата последнего бэкапа
    import glob as _glob
    backup_files = sorted(_glob.glob("/var/www/crm/GCRM-2/backups/*.sql.gz"))
    last_backup = None
    if backup_files:
        mtime = os.path.getmtime(backup_files[-1])
        last_backup = datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M")

    return {
        "db": {
            "deals": db.query(Deal).count(),
            "contacts": db.query(Contact).count(),
            "active_tasks": active_tasks,
            "overdue_tasks": overdue,
            "active_deals": active_deals,
            "last_backup": last_backup,
        },
        "cache": {"keys": cache_keys, "count": len(cache_keys), "ttl": _cache._ttl},
        "system": system_payload,
        "tg_configured": bool(os.getenv("TELEGRAM_BOT_TOKEN") and os.getenv("TELEGRAM_CHAT_ID")),
        "tg2_configured": bool(os.getenv("TELEGRAM_BOT2_TOKEN")),
        "tg3_configured": bool(os.getenv("TELEGRAM_ASSISTANT_BOT_TOKEN")),
    }

@app.post("/api/service/cache/clear")
def service_clear_cache(user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    _cache._data.clear()
    return {"ok": True, "message": "Кэш полностью очищен"}

# Расписание авто-отчёта (in-memory, сбрасывается при рестарте)
_bot_schedule = {"enabled": False, "time": "18:00"}

@app.get("/api/service/bot/schedule")
def get_bot_schedule(user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    return _bot_schedule

@app.post("/api/service/bot/schedule")
async def set_bot_schedule(payload: dict, user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    import re as _re
    t = payload.get("time", "18:00")
    if not _re.match(r"^\d{1,2}:\d{2}$", t):
        raise HTTPException(400, "Неверный формат времени (ЧЧ:ММ)")
    _bot_schedule["enabled"] = bool(payload.get("enabled", False))
    _bot_schedule["time"] = t
    return {"ok": True, "enabled": _bot_schedule["enabled"], "time": _bot_schedule["time"]}

@app.get("/api/service/run-repeats")
def service_run_repeats(user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    try:
        _process_repeat_deals()
        return {"ok": True, "message": "Проверка повторных сделок выполнена"}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/service/bot/report")
async def service_send_report(db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat  = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat:
        raise HTTPException(400, "TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID не заданы в .env")

    today = datetime.utcnow().date()
    lines = []

    # ── Заголовок
    lines.append(f"🌿 GrassCRM — отчёт за {today.strftime('%d.%m.%Y')}")
    lines.append("")

    # ── Сделки в работе
    active_deals = (
        db.query(Deal)
        .join(Stage)
        .filter(Stage.is_final == False)
        .options(joinedload(Deal.contact), joinedload(Deal.stage))
        .order_by(Stage.order, Deal.created_at.desc())
        .all()
    )
    if active_deals:
        lines.append("📋 Сделки в работе:")
        current_stage = None
        for deal in active_deals:
            stage_name = deal.stage.name if deal.stage else "Без стадии"
            if stage_name != current_stage:
                current_stage = stage_name
                lines.append(f"\n  ▸ {stage_name}")
            client = deal.contact.name if deal.contact else "Без клиента"
            total = f"{int(deal.total or 0):,}".replace(",", " ")
            lines.append(f"    · {deal.title} ({client}) — {total} руб.")
            if deal.deal_date:
                lines.append(f"      📅 {deal.deal_date.strftime('%d.%m.%Y %H:%M')}")
            if deal.address:
                lines.append(f"      📍 {deal.address}")
    else:
        lines.append("📋 Активных сделок нет.")

    lines.append("")

    # ── Задачи
    today_tasks = (
        db.query(Task)
        .filter(Task.is_done == False, Task.due_date == today)
        .all()
    )
    overdue_tasks = (
        db.query(Task)
        .filter(Task.is_done == False, Task.due_date < today)
        .order_by(Task.due_date)
        .all()
    )

    if today_tasks:
        lines.append("✅ Задачи на сегодня:")
        for t in today_tasks:
            lines.append(f"    · {t.title}")
        lines.append("")

    if overdue_tasks:
        lines.append(f"⚠️ Просрочено ({len(overdue_tasks)}):")
        for t in overdue_tasks[:5]:
            due = t.due_date.strftime("%d.%m") if t.due_date else "—"
            lines.append(f"    · {t.title} (до {due})")
        if len(overdue_tasks) > 5:
            lines.append(f"    ...и ещё {len(overdue_tasks) - 5}")
        lines.append("")

    if not today_tasks and not overdue_tasks:
        lines.append("✅ Задач на сегодня нет.")
        lines.append("")

    # ── Цитата
    try:
        row = db.execute(text("SELECT phrase FROM daily_phrases ORDER BY RANDOM() LIMIT 1")).fetchone()
        if row:
            phrase = row[0].replace("\\n", "\n")
            lines.append("· · · · · · · · · ·")
            lines.append(phrase)
    except Exception:
        pass

    msg_text = "\n".join(lines)
    async with httpx.AsyncClient(timeout=10) as client:
        r = await client.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat, "text": msg_text}
        )
        if not r.is_success:
            raise HTTPException(500, f"Telegram ошибка {r.status_code}: {r.text[:200]}")
    return {"ok": True, "message": "Отчёт отправлен в Telegram"}

@app.post("/api/service/backup")
def service_backup(user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    import subprocess as _sp, glob as _glob, re as _re
    db_url = os.getenv("DATABASE_URL", "")
    m = _re.match(r"postgresql(?:\+\w+)?://([^:]+):([^@]+)@([^:/]+):?(\d+)?/(\S+)", db_url)
    if not m:
        raise HTTPException(500, "Не удалось разобрать DATABASE_URL")
    db_user, db_pass, db_host, db_port, db_name = m.group(1), m.group(2), m.group(3), m.group(4) or "5432", m.group(5)
    db_name = db_name.split("?")[0]

    backup_dir = "/var/www/crm/GCRM-2/backups"
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_file = f"{backup_dir}/backup_{ts}.sql.gz"

    # Ищем pg_dump подходящей версии (17 → 16 → системный)
    pg_dump_cmd = None
    for candidate in ["/usr/lib/postgresql/17/bin/pg_dump", "/usr/lib/postgresql/16/bin/pg_dump", "pg_dump"]:
        try:
            r = _sp.run([candidate, "--version"], capture_output=True, text=True, timeout=3)
            if r.returncode == 0:
                pg_dump_cmd = candidate
                break
        except FileNotFoundError:
            continue
    if not pg_dump_cmd:
        raise HTTPException(500, "pg_dump не найден. Установите: apt install postgresql-client-17")

    env = {**os.environ, "PGPASSWORD": db_pass}
    try:
        dump = _sp.Popen(
            [pg_dump_cmd, "-h", db_host, "-p", db_port, "-U", db_user, "-d", db_name,
             "--no-password", "-F", "p"],
            stdout=_sp.PIPE, stderr=_sp.PIPE, env=env
        )
        with open(out_file, "wb") as f_out:
            gzip_proc = _sp.Popen(["gzip", "-c"], stdin=dump.stdout, stdout=f_out, stderr=_sp.PIPE)
        dump.stdout.close()
        _, dump_err = dump.communicate(timeout=120)
        gzip_proc.communicate(timeout=30)
        if dump.returncode != 0:
            if os.path.exists(out_file): os.remove(out_file)
            raise HTTPException(500, f"pg_dump ошибка: {dump_err.decode()[:300]}")
    except _sp.TimeoutExpired:
        raise HTTPException(500, "pg_dump timeout (120s)")

    # Оставляем последние 10 бэкапов
    all_backups = sorted(_glob.glob(f"{backup_dir}/backup_*.sql.gz"))
    for old in all_backups[:-10]:
        try: os.remove(old)
        except: pass

    size_kb = round(os.path.getsize(out_file) / 1024, 1)
    return {"ok": True, "message": f"Бэкап создан: backup_{ts}.sql.gz ({size_kb} КБ)"}


@app.post("/api/service/db/check")
def service_db_check(db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    try:
        ver = db.execute(text("SELECT version()")).scalar()
        return {"ok": True, "message": "БД доступна", "detail": str(ver)[:80]}
    except Exception as e:
        raise HTTPException(500, f"Ошибка: {e}")

@app.get("/api/service/logs")
def service_get_logs(n: int = 60, user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    import subprocess as _sp
    try:
        r = _sp.run(["journalctl", "-u", "crm.service", f"-n{n}", "--no-pager", "--output=short"],
                    capture_output=True, text=True, timeout=6)
        return {"ok": True, "logs": r.stdout}
    except Exception as e:
        return {"ok": False, "logs": str(e)}

def build_crm_context(db, req_year: int) -> dict:
    """
    Собирает полный контекст CRM для AI.
    Возвращает dict с deals, tasks, expenses, contacts, stats.
    """
    from datetime import date as _date

    today = _date.today()
    month_prefix = today.strftime("%Y-%m")

    # ── Стадии ────────────────────────────────────────────────────────────────
    stages = {s.id: s for s in db.query(Stage).all()}
    won_ids = {s.id for s in stages.values() if s.is_final and "успешно" in (s.name or "").lower()}

    # ── Сделки за год ─────────────────────────────────────────────────────────
    deals_q = db.query(Deal).filter(extract('year', Deal.created_at) == req_year)\
                .order_by(Deal.created_at.desc()).all()
    won_deals = [d for d in deals_q if d.stage_id in won_ids]
    active_deals = [d for d in deals_q if d.stage_id not in won_ids]

    total_rev = sum(d.total or 0 for d in won_deals)
    win_rate = round(len(won_deals) / len(deals_q) * 100, 1) if deals_q else 0

    # Последние 10 сделок для детального контекста
    recent_deals = []
    for d in deals_q[:10]:
        stage_name = stages[d.stage_id].name if d.stage_id and d.stage_id in stages else "—"
        contact_name = d.contact.name if d.contact else "—"
        recent_deals.append({
            "id": d.id,
            "title": d.title,
            "contact": contact_name,
            "stage": stage_name,
            "total": d.total or 0,
            "address": d.address or "",
            "created_at": d.created_at.strftime("%Y-%m-%d") if d.created_at else "",
        })

    # ── Расходы ───────────────────────────────────────────────────────────────
    expenses_q = db.query(Expense).filter(extract('year', Expense.date) == req_year)\
                   .order_by(Expense.date.desc()).all()
    total_exp = sum(e.amount or 0 for e in expenses_q)
    profit = total_rev - total_exp

    month_exp = sum(e.amount or 0 for e in expenses_q
                    if str(e.date)[:7] == month_prefix)

    recent_expenses = []
    for e in expenses_q[:10]:
        cat_name = e.category.name if e.category else "—"
        recent_expenses.append({
            "name": e.name,
            "amount": e.amount or 0,
            "date": str(e.date) if e.date else "",
            "category": cat_name,
        })

    # ── Задачи (все открытые) ─────────────────────────────────────────────────
    tasks_q = db.query(Task).filter(
        Task.status.notin_(["Завершена", "Выполнена"])
    ).order_by(Task.due_date.asc().nullslast()).limit(10).all()

    recent_tasks = []
    for t in tasks_q:
        recent_tasks.append({
            "title": t.title,
            "due_date": str(t.due_date) if t.due_date else "",
            "priority": t.priority or "Обычный",
            "status": t.status or "Открыта",
        })

    # Просроченные задачи
    overdue_count = db.query(Task).filter(
        Task.due_date < today,
        Task.status.notin_(["Завершена", "Выполнена"])
    ).count()

    # ── Контакты (топ-10 по активности) ──────────────────────────────────────
    contacts_q = db.query(Contact).order_by(Contact.id.desc()).limit(10).all()
    recent_contacts = []
    for c in contacts_q:
        recent_contacts.append({
            "name": c.name,
            "phone": c.phone or "",
            "settlement": c.settlement or "",
            "plot_area": c.plot_area,
        })

    # ── Итоговый словарь ──────────────────────────────────────────────────────
    return {
        "stats": {
            "year": req_year,
            "total_deals": len(deals_q),
            "won_deals": len(won_deals),
            "active_deals": len(active_deals),
            "win_rate_pct": win_rate,
            "revenue": round(total_rev, 2),
            "expenses_year": round(total_exp, 2),
            "expenses_month": round(month_exp, 2),
            "profit": round(profit, 2),
            "open_tasks": len(recent_tasks),
            "overdue_tasks": overdue_count,
        },
        "deals": recent_deals,
        "expenses": recent_expenses,
        "tasks": recent_tasks,
        "contacts": recent_contacts,
    }


@app.post("/api/service/ai/ask")
async def service_ai_ask(payload: ServiceAIAgentRequest, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    if not is_admin(user):
        raise HTTPException(403, "Admin only")

    base_url = (payload.base_url or OPENAI_BASE_URL or "").strip()
    access_id = (payload.access_id or OPENAI_ACCESS_ID or "").strip()
    model = (payload.model or OPENAI_MODEL or "gpt-4o-mini").strip()
    if not base_url:
        raise HTTPException(400, "Не задан OpenAI URL")
    if not access_id:
        raise HTTPException(400, "Не задан Access ID / API key")

    if not base_url.startswith("http://") and not base_url.startswith("https://"):
        base_url = f"https://{base_url}"

    req_year = payload.year or datetime.utcnow().year

    # ── Полный CRM-контекст ───────────────────────────────────────────────────
    crm = build_crm_context(db, req_year)
    stats = crm["stats"]

    # Форматируем в читаемый текст для промпта
    deals_text = "\n".join(
        f"  • #{d['id']} {d['title']} | {d['contact']} | {d['stage']} | {d['total']:,.0f}₽ | {d['created_at']}"
        for d in crm["deals"]
    ) or "  нет данных"

    tasks_text = "\n".join(
        f"  • [{t['priority']}] {t['title']} | срок: {t['due_date'] or '—'} | {t['status']}"
        for t in crm["tasks"]
    ) or "  нет данных"

    expenses_text = "\n".join(
        f"  • {e['name']} | {e['amount']:,.0f}₽ | {e['category']} | {e['date']}"
        for e in crm["expenses"]
    ) or "  нет данных"

    contacts_text = "\n".join(
        f"  • {c['name']}{' | ' + c['phone'] if c['phone'] else ''}"
        f"{' | ' + c['settlement'] if c['settlement'] else ''}"
        f"{' | ' + str(c['plot_area']) + ' сот' if c['plot_area'] else ''}"
        for c in crm["contacts"]
    ) or "  нет данных"

    context = f"""=== СТАТИСТИКА ({stats['year']}) ===
Сделок всего: {stats['total_deals']} | Успешных: {stats['won_deals']} ({stats['win_rate_pct']}%) | Активных: {stats['active_deals']}
Выручка: {stats['revenue']:,.0f}₽ | Расходы за год: {stats['expenses_year']:,.0f}₽ | Прибыль: {stats['profit']:,.0f}₽
Расходы за месяц: {stats['expenses_month']:,.0f}₽
Открытых задач: {stats['open_tasks']} | Просроченных: {stats['overdue_tasks']}

=== ПОСЛЕДНИЕ СДЕЛКИ (10) ===
{deals_text}

=== ОТКРЫТЫЕ ЗАДАЧИ (10) ===
{tasks_text}

=== ПОСЛЕДНИЕ РАСХОДЫ (10) ===
{expenses_text}

=== ПОСЛЕДНИЕ КЛИЕНТЫ (10) ===
{contacts_text}"""

    system_prompt = (
        "Ты AI-ассистент CRM для сервиса покоса и ландшафтных работ GrassCRM. "
        "Тебе передан полный контекст бизнеса: сделки, задачи, расходы, клиенты. "
        "Используй эти данные для точных и конкретных ответов. "
        "Отвечай строго на русском, коротко и по делу. "
        "Если даёшь рекомендации, разделяй их на: "
        "1) быстрые шаги на 7 дней, 2) системные шаги на месяц."
    )

    # ── Контекст конкретной записи (если открыта сделка/задача/расход) ─────────
    record_context = ""
    if payload.context_type and payload.context_id:
        try:
            ctype = payload.context_type.lower()
            cid   = payload.context_id
            if ctype == "deal":
                rec = db.query(Deal).filter(Deal.id == cid).first()
                if rec:
                    stage_name = rec.stage.name if rec.stage else "—"
                    contact_name = rec.contact.name if rec.contact else "—"
                    record_context = (
                        f"\n=== ТЕКУЩАЯ ЗАПИСЬ: СДЕЛКА #{cid} ===\n"
                        f"Название: {rec.title}\n"
                        f"Клиент: {contact_name}\n"
                        f"Стадия: {stage_name}\n"
                        f"Сумма: {rec.total or 0:,.0f}₽\n"
                        f"Адрес: {rec.address or '—'}\n"
                        f"Заметки: {(rec.notes or '')[:300]}\n"
                    )
            elif ctype == "task":
                rec = db.query(Task).filter(Task.id == cid).first()
                if rec:
                    record_context = (
                        f"\n=== ТЕКУЩАЯ ЗАПИСЬ: ЗАДАЧА #{cid} ===\n"
                        f"Название: {rec.title}\n"
                        f"Статус: {rec.status}\n"
                        f"Приоритет: {rec.priority}\n"
                        f"Срок: {rec.due_date or '—'}\n"
                        f"Описание: {(rec.description or '')[:300]}\n"
                    )
            elif ctype == "expense":
                rec = db.query(Expense).filter(Expense.id == cid).first()
                if rec:
                    cat_name = rec.category.name if rec.category else "—"
                    record_context = (
                        f"\n=== ТЕКУЩАЯ ЗАПИСЬ: РАСХОД #{cid} ===\n"
                        f"Название: {rec.name}\n"
                        f"Сумма: {rec.amount:,.0f}₽\n"
                        f"Дата: {rec.date}\n"
                        f"Категория: {cat_name}\n"
                    )
            elif ctype == "contact":
                rec = db.query(Contact).filter(Contact.id == cid).first()
                if rec:
                    record_context = (
                        f"\n=== ТЕКУЩАЯ ЗАПИСЬ: КЛИЕНТ #{cid} ===\n"
                        f"Имя: {rec.name}\n"
                        f"Телефон: {rec.phone or '—'}\n"
                        f"Посёлок: {rec.settlement or '—'}\n"
                        f"Площадь участка: {rec.plot_area or '—'} сот\n"
                    )
        except Exception as _e:
            pass  # не критично, продолжаем без контекста записи

    payload_json = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"Контекст CRM:\n{context}{record_context}\n\nЗапрос: {payload.prompt}"},
        ],
        "temperature": 0.4,
    }

    url = f"{base_url.rstrip('/')}/chat/completions"
    headers = {
        "Authorization": f"Bearer {access_id}",
        "Content-Type": "application/json",
    }

    try:
        async with httpx.AsyncClient(timeout=45) as client:
            r = await client.post(url, json=payload_json, headers=headers)
    except Exception as e:
        raise HTTPException(502, f"Ошибка подключения к AI API: {e}")

    if not r.is_success:
        raise HTTPException(502, f"AI API {r.status_code}: {r.text[:300]}")

    data = r.json()
    text_out = ((data.get("choices") or [{}])[0].get("message") or {}).get("content", "").strip()
    if not text_out:
        raise HTTPException(502, "AI API вернул пустой ответ")

    # Логируем использование токенов
    usage = data.get("usage") or {}
    if usage:
        try:
            log = AiUsageLog(
                prompt_tokens=usage.get("prompt_tokens", 0),
                completion_tokens=usage.get("completion_tokens", 0),
                total_tokens=usage.get("total_tokens", 0),
                source="crm"
            )
            db.add(log)
            db.commit()
        except Exception:
            db.rollback()

    return {"ok": True, "answer": text_out, "model": model, "year": req_year}


@app.post("/api/service/ai/action")
async def service_ai_action(
    payload: AIActionRequest,
    db: DBSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """
    Универсальный executor AI-действий.
    Принимает action + data, выполняет операцию в БД, возвращает результат.
    Вызывается из CRM-интерфейса (AI-кнопки) и из bot.py.

    Поддерживаемые action:
      create_task    — создать задачу
      create_deal    — создать сделку
      create_expense — создать расход
      change_status  — сменить стадию сделки (deal_id + stage_name или stage_id)
      add_comment    — добавить комментарий к сделке (deal_id + text)
    """
    if not is_admin(user):
        raise HTTPException(403, "Admin only")

    action = (payload.action or "").strip()
    data   = payload.data or {}

    # ── create_task ───────────────────────────────────────────────────────────
    if action == "create_task":
        title = (data.get("title") or "").strip()
        if not title:
            raise HTTPException(400, "Поле title обязательно")

        contact_id = data.get("contact_id")
        if not contact_id and data.get("contact_name"):
            c = db.query(Contact).filter(
                Contact.name.ilike(f"%{data['contact_name']}%")
            ).first()
            contact_id = c.id if c else None

        due = None
        if data.get("due_date"):
            try:
                from datetime import date as _d
                due = _d.fromisoformat(str(data["due_date"])[:10])
            except Exception:
                pass

        task = Task(
            title=title,
            description=data.get("description"),
            due_date=due,
            priority=data.get("priority") or "Обычный",
            status="Открыта",
            contact_id=contact_id,
            deal_id=data.get("deal_id"),
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        _cache.invalidate("tasks")
        return {"ok": True, "action": action, "id": task.id, "title": task.title}

    # ── create_expense ────────────────────────────────────────────────────────
    if action == "create_expense":
        name   = (data.get("name") or "").strip()
        amount = data.get("amount")
        if not name or amount is None:
            raise HTTPException(400, "Поля name и amount обязательны")

        from datetime import date as _d
        exp_date = _d.today()
        if data.get("date"):
            try:
                exp_date = _d.fromisoformat(str(data["date"])[:10])
            except Exception:
                pass

        cat_name = (data.get("category") or "Прочее").strip()
        cat = db.query(ExpenseCategory).filter(
            ExpenseCategory.name.ilike(cat_name)
        ).first()
        if not cat:
            cat = db.query(ExpenseCategory).first()

        expense = Expense(
            name=name,
            amount=float(amount),
            date=exp_date,
            category_id=cat.id if cat else None,
        )
        db.add(expense)
        db.commit()
        db.refresh(expense)
        _cache.invalidate("expenses")
        return {"ok": True, "action": action, "id": expense.id, "name": expense.name}

    # ── create_deal ───────────────────────────────────────────────────────────
    if action == "create_deal":
        title = (data.get("title") or "").strip()
        if not title:
            raise HTTPException(400, "Поле title обязательно")

        # Найти или создать контакт
        contact_id = data.get("contact_id")
        if not contact_id:
            contact_name = (data.get("contact_name") or "Неизвестный клиент").strip()
            phone        = data.get("phone")
            contact = None
            if phone:
                contact = db.query(Contact).filter(Contact.phone == phone).first()
            if not contact:
                contact = db.query(Contact).filter(
                    Contact.name.ilike(f"%{contact_name}%")
                ).first()
            if not contact:
                contact = Contact(name=contact_name, phone=phone)
                db.add(contact)
                db.flush()
                db.refresh(contact)
            contact_id = contact.id

        # Первая стадия
        first_stage = db.query(Stage).order_by(Stage.order.asc()).first()

        deal = Deal(
            title=title,
            contact_id=contact_id,
            stage_id=first_stage.id if first_stage else None,
            notes=data.get("notes") or "",
            address=data.get("address") or "",
        )
        db.add(deal)
        db.commit()
        db.refresh(deal)
        _cache.invalidate("deals", "years")
        return {"ok": True, "action": action, "id": deal.id, "title": deal.title}

    # ── change_status ─────────────────────────────────────────────────────────
    if action == "change_status":
        deal_id = data.get("deal_id")
        if not deal_id:
            raise HTTPException(400, "Поле deal_id обязательно")

        deal = db.query(Deal).filter(Deal.id == deal_id).first()
        if not deal:
            raise HTTPException(404, f"Сделка #{deal_id} не найдена")

        # Ищем стадию по id или по имени
        stage = None
        if data.get("stage_id"):
            stage = db.query(Stage).filter(Stage.id == data["stage_id"]).first()
        elif data.get("stage_name"):
            stage = db.query(Stage).filter(
                Stage.name.ilike(f"%{data['stage_name']}%")
            ).first()

        if not stage:
            available = [s.name for s in db.query(Stage).order_by(Stage.order).all()]
            raise HTTPException(404, f"Стадия не найдена. Доступные: {available}")

        old_stage = deal.stage.name if deal.stage else "—"
        deal.stage_id = stage.id

        # Если финальная успешная — закрываем
        if stage.is_final and "успешно" in (stage.name or "").lower():
            deal.closed_at = datetime.utcnow()

        db.commit()
        _cache.invalidate("deals", "years")
        return {
            "ok": True, "action": action,
            "deal_id": deal_id,
            "from": old_stage,
            "to": stage.name,
        }

    # ── add_comment ───────────────────────────────────────────────────────────
    if action == "add_comment":
        deal_id = data.get("deal_id")
        text    = (data.get("text") or "").strip()
        if not deal_id or not text:
            raise HTTPException(400, "Поля deal_id и text обязательны")

        deal = db.query(Deal).filter(Deal.id == deal_id).first()
        if not deal:
            raise HTTPException(404, f"Сделка #{deal_id} не найдена")

        comment = DealComment(
            deal_id=deal_id,
            text=text,
            user_name=user.get("name") or "AI",
            user_id=user.get("sub"),
        )
        db.add(comment)
        db.commit()
        db.refresh(comment)
        return {
            "ok": True, "action": action,
            "deal_id": deal_id,
            "comment_id": comment.id,
        }

    raise HTTPException(400, f"Неизвестный action: '{action}'. "
                            f"Доступные: create_task, create_deal, create_expense, "
                            f"change_status, add_comment")


# ══════════════════════════════════════════════════════
#  🧠  AI MEMORY ENDPOINTS
# ══════════════════════════════════════════════════════

class AiMemorySaveRequest(BaseModel):
    chat_id:     str
    role:        str = "user"        # user | assistant | fact
    content:     str
    memory_type: str = "message"     # message | fact | preference
    importance:  int = 1
    metadata:    dict = {}
    ttl_days:    Optional[int] = None  # сколько дней хранить (None = вечно)


@app.post("/api/service/ai/memory", status_code=201)
def ai_memory_save(
    payload: AiMemorySaveRequest,
    db: DBSession = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    """Сохранить запись в память AI (вызывается из бота)."""
    from datetime import timedelta as _td

    expires = None
    if payload.ttl_days:
        expires = datetime.utcnow() + _td(days=payload.ttl_days)

    # Для сообщений — не дублируем если уже есть точно такое же за последние 5 мин
    if payload.memory_type == "message":
        from datetime import timedelta as _td2
        cutoff = datetime.utcnow() - _td2(minutes=5)
        exists = db.query(AiMemory).filter(
            AiMemory.chat_id == str(payload.chat_id),
            AiMemory.content == payload.content,
            AiMemory.created_at >= cutoff,
        ).first()
        if exists:
            return {"ok": True, "id": exists.id, "duplicate": True}

    mem = AiMemory(
        chat_id     = str(payload.chat_id),
        role        = payload.role,
        content     = payload.content[:2000],
        memory_type = payload.memory_type,
        importance  = payload.importance,
        metadata_   = payload.metadata,
        expires_at  = expires,
    )
    db.add(mem)
    db.commit()
    db.refresh(mem)
    return {"ok": True, "id": mem.id}


@app.get("/api/service/ai/memory")
def ai_memory_search(
    chat_id: str,
    q: Optional[str] = None,
    memory_type: Optional[str] = None,
    limit: int = 20,
    db: DBSession = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    """
    Получить релевантные воспоминания для chat_id.
    Если передан q — ищет по вхождению текста (ILIKE).
    Всегда возвращает: факты (facts) + последние сообщения.
    """
    from datetime import datetime as _dt

    now = _dt.utcnow()
    base_q = db.query(AiMemory).filter(
        AiMemory.chat_id == str(chat_id),
    ).filter(
        (AiMemory.expires_at == None) | (AiMemory.expires_at > now)
    )

    if memory_type:
        base_q = base_q.filter(AiMemory.memory_type == memory_type)

    # Факты и предпочтения — всегда тянем (важные первыми)
    facts = base_q.filter(
        AiMemory.memory_type.in_(["fact", "preference"])
    ).order_by(AiMemory.importance.desc(), AiMemory.created_at.desc()).limit(30).all()

    # Последние сообщения диалога
    recent_msgs = base_q.filter(
        AiMemory.memory_type == "message"
    ).order_by(AiMemory.created_at.desc()).limit(10).all()

    # Текстовый поиск если задан запрос
    search_results = []
    if q and len(q.strip()) >= 3:
        search_results = base_q.filter(
            AiMemory.content.ilike(f"%{q.strip()}%")
        ).order_by(AiMemory.importance.desc(), AiMemory.created_at.desc()).limit(10).all()

    def _fmt(m: AiMemory) -> dict:
        return {
            "id":          m.id,
            "role":        m.role,
            "content":     m.content,
            "memory_type": m.memory_type,
            "importance":  m.importance,
            "created_at":  m.created_at.isoformat() if m.created_at else None,
            "metadata":    m.metadata_ or {},
        }

    # Дедупликация по id
    seen = set()
    merged = []
    for m in list(facts) + list(search_results) + list(reversed(recent_msgs)):
        if m.id not in seen:
            seen.add(m.id)
            merged.append(_fmt(m))

    return {"ok": True, "count": len(merged), "memories": merged}


@app.delete("/api/service/ai/memory")
def ai_memory_clear(
    chat_id: str,
    memory_type: Optional[str] = None,
    db: DBSession = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    """Очистить память для chat_id (опционально — только определённый тип)."""
    q = db.query(AiMemory).filter(AiMemory.chat_id == str(chat_id))
    if memory_type:
        q = q.filter(AiMemory.memory_type == memory_type)
    deleted = q.delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": deleted}


@app.get("/api/service/ai/usage")
def service_ai_usage(db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    """Статистика использования токенов AI агента."""
    if not is_admin(user): raise HTTPException(403, "Admin only")
    quota = int(os.getenv("AI_TOKEN_QUOTA", "500000"))
    from sqlalchemy import func as _func
    total = db.execute(text("SELECT COALESCE(SUM(total_tokens),0) FROM ai_usage_log")).scalar() or 0
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today = db.execute(text("SELECT COALESCE(SUM(total_tokens),0) FROM ai_usage_log WHERE created_at >= :ts"), {"ts": today_start}).scalar() or 0
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month = db.execute(text("SELECT COALESCE(SUM(total_tokens),0) FROM ai_usage_log WHERE created_at >= :ts"), {"ts": month_start}).scalar() or 0
    calls_total = db.execute(text("SELECT COUNT(*) FROM ai_usage_log")).scalar() or 0
    calls_month = db.execute(text("SELECT COUNT(*) FROM ai_usage_log WHERE created_at >= :ts"), {"ts": month_start}).scalar() or 0
    return {
        "quota": quota,
        "used_total": int(total),
        "used_month": int(month),
        "used_today": int(today),
        "remaining": max(0, quota - int(total)),
        "pct": round(int(total) / quota * 100, 1) if quota else 0,
        "calls_total": int(calls_total),
        "calls_month": int(calls_month),
    }


@app.post("/api/service/restart")
def service_restart(user: dict = Depends(get_current_user)):
    if not is_admin(user): raise HTTPException(403, "Admin only")
    import subprocess as _sp, threading as _th
    def _do():
        import time as _t; _t.sleep(2)
        _sp.Popen(["systemctl", "restart", "crm"])
    _th.Thread(target=_do, daemon=True).start()
    return {"ok": True, "message": "Перезапуск через 2 секунды…"}

# ── NOTES ─────────────────────────────────────────────────────────────────────
import uuid as _uuid

def _note_to_dict(n: Note) -> dict:
    return {
        "id": n.id,
        "title": n.title or "",
        "body": n.body or "",
        "color": n.color or "",
        "pinned": n.pinned or False,
        "label": n.label or "",
        "checklist": n.checklist or [],
        "created": int(n.created_at.timestamp() * 1000) if n.created_at else 0,
        "updated": int(n.updated_at.timestamp() * 1000) if n.updated_at else 0,
    }

@app.get("/api/notes")
def get_notes(db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    notes_list = (db.query(Note)
                  .filter(Note.user_id == user["sub"])
                  .order_by(Note.updated_at.desc())
                  .all())
    return [_note_to_dict(n) for n in notes_list]

@app.post("/api/notes", status_code=201)
def create_note(data: NoteCreate, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    note = Note(
        id=data.id or str(_uuid.uuid4()),
        user_id=user["sub"],
        title=data.title,
        body=data.body,
        color=data.color,
        pinned=data.pinned,
        label=data.label,
        checklist=data.checklist,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(note)
    db.commit()
    db.refresh(note)
    return _note_to_dict(note)

@app.patch("/api/notes/{note_id}")
def update_note(note_id: str, data: NoteUpdate, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    note = db.query(Note).filter(Note.id == note_id, Note.user_id == user["sub"]).first()
    if not note:
        raise HTTPException(404, "Заметка не найдена")
    if data.title is not None: note.title = data.title
    if data.body is not None: note.body = data.body
    if data.color is not None: note.color = data.color
    if data.pinned is not None: note.pinned = data.pinned
    if data.label is not None: note.label = data.label
    if data.checklist is not None: note.checklist = data.checklist
    note.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(note)
    return _note_to_dict(note)

@app.delete("/api/notes/{note_id}", status_code=204)
def delete_note(note_id: str, db: DBSession = Depends(get_db), user: dict = Depends(get_current_user)):
    note = db.query(Note).filter(Note.id == note_id, Note.user_id == user["sub"]).first()
    if not note:
        raise HTTPException(404, "Заметка не найдена")
    db.delete(note)
    db.commit()

@app.get("/api/bot-faq")
def get_bot_faq(db: DBSession = Depends(get_db), _=Depends(get_current_user)):
    items = (db.query(BotFaq)
             .filter(BotFaq.active == True)
             .order_by(BotFaq.priority.desc())
             .all())
    return [
        {"id": i.id, "intent": i.intent, "question_example": i.question_example, "answer": i.answer}
        for i in items
    ]

@app.get("/{full_path:path}", response_class=FileResponse)
async def serve_frontend(full_path: str):
    path = f"./{full_path.strip()}" if full_path else "./index.html"
    return FileResponse(path if os.path.isfile(path) else "./index.html")

print(f"main.py (v14.0) loaded — backup, cpu fix, analytics.", flush=True)
